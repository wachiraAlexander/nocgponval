"""
================================================================================
  REPORT GENERATOR  -  GPON  &  ENTERPRISE
================================================================================

Single-file desktop app for generating two telecom reports from a daily
Issue.xlsx export:

    Tab 1: GPON       -> "<Date> GPON Report.xlsx"
    Tab 2: ENTERPRISE -> "<Date> Enterprise Report.xlsx"

Both tabs share the same input file but produce independent outputs. Both
share the same visual treatment so the two reports look like a matched pair.

USAGE
    pip install pandas openpyxl
    pip install tkinterdnd2     # optional, enables drag-and-drop
    python app.py

CONTENTS  (search for the banner you want)
    SECTION 1   Imports + module-level setup
    SECTION 2   Constants, palette, regex patterns
    SECTION 3   Assignee name cleaning (raw email -> "Frank Gitiria")
    SECTION 4   Generic helpers (column lookup, drop columns, SLA)
    SECTION 5   Styling primitives (title bar, header band, SLA tiers,
                row shading, mini-summary)
    SECTION 6   Adaptive Summary layout (brick-wall packing)
    SECTION 7   ENTERPRISE pipeline
    SECTION 8   GPON pipeline
    SECTION 9   GUI (notebook with two tabs, settings, recent files,
                live preview, bucket chart, drag-and-drop)
    SECTION 10  Entry point

Authors: Emmanuel Mutua & Alex Wachira
"""

# ============================================================================
# SECTION 1   Imports + module-level setup
# ============================================================================
from __future__ import annotations

import ast
import json
import os
import re
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd

# tkinter imports are deferred to when the GUI launches, so the file can be
# imported and used as a library on systems without a display (headless servers,
# CI pipelines, etc.)
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    _TK_AVAILABLE = True
except ImportError:
    tk = None
    ttk = None
    filedialog = None
    messagebox = None
    scrolledtext = None
    _TK_AVAILABLE = False

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False
    TkinterDnD = None
    DND_FILES = None


# When tkinter is missing (e.g. headless server), stub out the names that
# class definitions reference so the module still imports as a library.
class _TkStub:
    def __getattr__(self, _):  # any attribute access returns object (a base class)
        return object
    def __call__(self, *a, **kw):
        return self

if not _TK_AVAILABLE:
    tk = _TkStub()
    ttk = _TkStub()
    filedialog = _TkStub()
    messagebox = _TkStub()
    scrolledtext = _TkStub()

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# ============================================================================
# SECTION 2   Constants, palette, regex patterns
# ============================================================================

APP_VERSION = "2.0.0"
APP_AUTHOR = "Emmanuel Mutua & Alex Wachira"

SETTINGS_FILE = Path(__file__).resolve().parent / "report_settings.json"

DEFAULT_SETTINGS: Dict[str, object] = {
    "recent_gpon_files": [],
    "recent_ent_files": [],
    "sla_green_max": 1,
    "sla_yellow_max": 3,
    "sla_orange_max": 7,
    "sla_red_max": 14,
    "theme": "clam",
    "last_output_dir": "",
}

# ----- Input data conventions (column names, normalized to lowercase) -----
DROP_COLUMNS = {"sr", "docstatus"}
STATUS_COLUMN_NORMALIZED = "status"
STATUS_TO_EXCLUDE = "closed"
CATEGORY1_COLUMN_NORMALIZED = "category 1"
CATEGORY2_COLUMN_NORMALIZED = "category 2"
CATEGORY3_COLUMN_NORMALIZED = "category 3"
CATEGORY4_COLUMN_NORMALIZED = "category 4"
ASSIGN_COLUMN_NORMALIZED = "_assign"
CREATION_COLUMN_NORMALIZED = "creation"
SUBJECT_COLUMN_NORMALIZED = "subject"
DESCRIPTION_COLUMN_NORMALIZED = "description"
SUBSCRIPTION_CANDIDATES = ["subscription", "subscription id"]
CUSTOMER_CANDIDATES = ["customer", "customer name"]
GPON_KEYWORD = "gpon"

STATUS_SPECIAL_SHEETS: Dict[str, str] = {
    "customer action": "Customer Action",
    "under monitoring": "Under Monitoring",
    "temporary restoration": "Temporary Restoration",
}

# ----- Color palette  -----
PASTEL_HEX_COLORS: List[str] = [
    "E8F0FE", "E6F4EA", "FFF3E0", "F3E5F5", "E0F2F1",
    "FFFDE7", "EDE7F6", "F1F8E9", "FCE4EC", "E0E0E0",
    "E3F2FD", "F1F8FF", "E8EAF6", "F9FBE7", "E0F7FA",
]

HEADER_FILL_HEX = "1F3864"
HEADER_FONT_HEX = "FFFFFF"
TITLE_BAR_FILL_HEX = "2E5597"
TITLE_BAR_FONT_HEX = "FFFFFF"
KPI_FILL_HEX = "F2F2F2"
KPI_ACCENT_HEX = "1F3864"
GRID_BORDER_HEX = "BFBFBF"
SUB_HEADER_FILL_HEX = "D9E1F2"

SLA_GREEN_HEX = "C6EFCE"
SLA_YELLOW_HEX = "FFEB9C"
SLA_ORANGE_HEX = "FFC79F"
SLA_RED_HEX = "FFB2B2"
SLA_DARKRED_HEX = "C00000"

TAB_COLORS: Dict[str, str] = {
    # Enterprise
    "Summary":                          "FFD700",
    "BB LOS":                           "4472C4",
    "BB on Extreme Low RX":             "4472C4",
    "BB on Low RX":                     "4472C4",
    "BB on CRC Errors":                 "4472C4",
    "6150|6120|1050-LOS":               "ED7D31",
    "6150|6120|1050-Extreme Low RX":    "ED7D31",
    "6150|6120|1050-Low RX":            "ED7D31",
    "6150|6120|1050-CRC":               "ED7D31",
    "Access Ring-LOS":                  "70AD47",
    "Access Ring-Low RX":               "70AD47",
    "Access Ring-Offline":              "70AD47",
    "Equipment TTs":                    "7030A0",
    "Power":                            "C00000",
    "CESR":                             "808080",
    # GPON
    "GPON SUMMARY":                     "FFD700",
    "GPON LOS Fiber Cut":               "4472C4",
    "GPON Extreme Low RX":              "4472C4",
    "GPON Low RX":                      "4472C4",
    "GPON LOSi":                        "4472C4",
    "GPON SFi":                         "4472C4",
    "GPON LOFi":                        "4472C4",
    "Customer Action":                  "ED7D31",
    "Under Monitoring":                 "70AD47",
    "Temporary Restoration":            "4472C4",
    "Double Tickets":                   "C00000",
    "GPON LOS - Customer Action":       "ED7D31",
    "MDU - LOS":                        "7030A0",
    "MDU - Low RX":                     "7030A0",
}
TAB_COLOR_GPON_CATEGORY = "4472C4"
TAB_COLOR_SPLITTER = "70AD47"
TAB_COLOR_FALLBACK = "808080"


def tab_color_for(sheet_name: str) -> str:
    """Hex (no '#') for a sheet's tab. Exact match first, then family rules."""
    if sheet_name in TAB_COLORS:
        return TAB_COLORS[sheet_name]
    if sheet_name.startswith("GPON - "):
        return TAB_COLOR_GPON_CATEGORY
    if sheet_name.startswith("Splitter - ") or sheet_name == "Splitter":
        return TAB_COLOR_SPLITTER
    return TAB_COLOR_FALLBACK


# Priority order for displaying buckets in the Enterprise Summary.
# Most operationally important first — these set the eye's anchor.
ENTERPRISE_BUCKET_PRIORITY: List[str] = [
    # Backbone is always first — most service-affecting
    "BB LOS",
    "BB on Extreme Low RX",
    "BB on Low RX",
    "BB on CRC Errors",
    # RAN equipment (6150 / 6120 / 1050) next
    "6150|6120|1050-LOS",
    "6150|6120|1050-Extreme Low RX",
    "6150|6120|1050-Low RX",
    "6150|6120|1050-CRC",
    # Access ring after backbone+RAN
    "Access Ring-LOS",
    "Access Ring-Low RX",
    "Access Ring-Offline",
    # Power, equipment, CESR
    "Power",
    "Equipment TTs",
    "CESR",
]


# ============================================================================
# SECTION 3   Assignee name cleaning
# ============================================================================
#  Input cells look like:    ["frank.gitiria@jtl.co.ke"]
#                            ["a@x.com", "b@x.com"]
#                            wkiptoo@jtl.co.ke
#                            ""  /  None  /  NaN
#  Output we want:           "Frank Gitiria"
#                            "A, B"
#                            "Wkiptoo"
#                            "(Unassigned)"

_EMAIL_LOCAL_RE = re.compile(r"^([^@]+)@.+$")


def _split_local_to_words(local: str) -> str:
    """Convert email local-part into title-cased words.
    'frank.gitiria' -> 'Frank Gitiria'
    'sharon_kibet'  -> 'Sharon Kibet'
    """
    if not local:
        return ""
    s = local.replace("_", " ").replace("-", " ").replace(".", " ")
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)  # camelCase -> camel Case
    parts = [p for p in s.split() if p]
    return " ".join(p[:1].upper() + p[1:].lower() for p in parts)


def clean_assignee(raw: object) -> str:
    """Convert messy _Assign values into clean display names."""
    if raw is None:
        return "(Unassigned)"
    s = str(raw).strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return "(Unassigned)"

    emails: List[str] = []
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                emails = [str(x).strip() for x in parsed if str(x).strip()]
            elif isinstance(parsed, str):
                emails = [parsed.strip()]
        except (ValueError, SyntaxError):
            inner = s[1:-1].replace('"', "").replace("'", "")
            emails = [p.strip() for p in inner.split(",") if p.strip()]
    else:
        emails = [p.strip() for p in s.split(",") if p.strip()]

    if not emails:
        return "(Unassigned)"

    pretty: List[str] = []
    for e in emails:
        m = _EMAIL_LOCAL_RE.match(e)
        local = m.group(1) if m else e
        name = _split_local_to_words(local)
        if name:
            pretty.append(name)
    if not pretty:
        return "(Unassigned)"
    return ", ".join(pretty)


def clean_assignee_series(series: pd.Series) -> pd.Series:
    """Vectorized clean_assignee for a column."""
    if series is None or len(series) == 0:
        return pd.Series([], dtype=str)
    return series.apply(clean_assignee)


# ============================================================================
# SECTION 4   Generic helpers
# ============================================================================

def _normalized(name: object) -> str:
    return str(name).strip().lower()


def _find_column_case_insensitive(df: pd.DataFrame, target_norm: str) -> Optional[str]:
    for col in df.columns:
        if _normalized(col) == target_norm:
            return col
    return None


def _find_any_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for cand in candidates:
        col = _find_column_case_insensitive(df, cand)
        if col is not None:
            return col
    return None


def _drop_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    cols_to_drop: List[str] = []
    targets = {c.lower() for c in DROP_COLUMNS}
    for col in df.columns:
        if _normalized(col) in targets:
            cols_to_drop.append(col)
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    return df, cols_to_drop


def _filter_closed_status(df: pd.DataFrame) -> Tuple[pd.DataFrame, int, Optional[str]]:
    status_col = _find_column_case_insensitive(df, STATUS_COLUMN_NORMALIZED)
    if status_col is None:
        return df, 0, None
    norm = df[status_col].astype(str).str.strip().str.lower()
    mask = norm != STATUS_TO_EXCLUDE
    dropped = int((~mask).sum())
    df = df.loc[mask].reset_index(drop=True)
    return df, dropped, status_col


def _append_sla_days(df: pd.DataFrame) -> pd.DataFrame:
    """Append an SLA column = integer days since Creation."""
    if df is None or df.empty:
        return df
    out = df.copy()
    creation_col = _find_column_case_insensitive(out, CREATION_COLUMN_NORMALIZED)
    if creation_col is None:
        out["SLA"] = pd.Series([pd.NA] * len(out))
        return out
    try:
        creation_dt = pd.to_datetime(out[creation_col], errors="coerce")
        now = pd.Timestamp.now()
        sla_days = (now - creation_dt).dt.days
        try:
            out["SLA"] = sla_days.astype("Int64")
        except Exception:
            out["SLA"] = sla_days
    except Exception:
        out["SLA"] = pd.Series([pd.NA] * len(out))
    return out


def _norm_series(df: pd.DataFrame, col: Optional[str]) -> pd.Series:
    if df is None or df.empty:
        return pd.Series([], dtype=str)
    if col is None or col not in df.columns:
        return pd.Series([""] * len(df), index=df.index, dtype=str)
    return df[col].astype(str).fillna("").str.strip().str.lower()


def _excel_safe_sheet_name(base: str, used: set) -> str:
    """Excel-safe, unique sheet name (max 31 chars, no [ ] : * ? / \\)."""
    invalid = set('[]:*?/\\')
    name = "".join(ch for ch in str(base) if ch not in invalid).strip()
    if not name:
        name = "Sheet"
    name = name[:31]
    candidate = name
    i = 2
    while candidate in used or candidate == "":
        suffix = f" ({i})"
        base_trunc = name[: max(0, 31 - len(suffix))]
        candidate = f"{base_trunc}{suffix}" if base_trunc else f"Sheet{suffix}"
        i += 1
    used.add(candidate)
    return candidate


def _block_counts(df: pd.DataFrame, assign_col: Optional[str]) -> pd.DataFrame:
    """[_Assign, Count] aggregated and pretty-named, sorted desc."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["_Assign", "Count"]).astype({"Count": int})
    if assign_col and assign_col in df.columns:
        labels = clean_assignee_series(df[assign_col])
    else:
        labels = pd.Series(["(Unassigned)"] * len(df))
    out = (
        pd.DataFrame({"_Assign": labels})
        .groupby(["_Assign"], dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(["Count", "_Assign"], ascending=[False, True])
    )
    try:
        out["Count"] = out["Count"].astype(int)
    except Exception:
        pass
    return out


# ============================================================================
# SECTION 5   Styling primitives
# ============================================================================
#
# Layout convention for ALL data sheets:
#     row 1            Title bar (merged across data columns)
#     row 2            Column headers (navy band, white text)
#     row 3..end       Data rows (pastel-shaded by assignee, SLA tiered)
#     right columns    Mini Assignee/Count summary
#
# Convention for ALL summary sheets:
#     adaptive packing — see SECTION 6


def _thin_side(color: str = GRID_BORDER_HEX) -> Side:
    return Side(style="thin", color=color)


def _all_borders(color: str = GRID_BORDER_HEX) -> Border:
    s = _thin_side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def _autosize_columns(ws, df: pd.DataFrame, max_width: int = 60) -> None:
    """Pick sensible column widths from header + first 200 sample rows."""
    for j, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        try:
            sample = df[col_name].astype(str).head(200)
            for v in sample:
                if v and v != "nan":
                    max_len = max(max_len, len(v))
        except Exception:
            pass
        width = min(max_len + 2, max_width)
        width = max(width, 8)
        if str(col_name).strip().lower() == "sla":
            width = 7
        ws.column_dimensions[get_column_letter(j)].width = width


def _apply_header_styling(ws, header_row: int, n_cols: int) -> None:
    """Bold white-on-navy header band with medium bottom border."""
    fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    font = Font(bold=True, color=HEADER_FONT_HEX, size=11)
    align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    bottom_border = Border(bottom=Side(style="medium", color="000000"))
    ws.row_dimensions[header_row].height = 22
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = bottom_border


def _apply_title_bar(ws, sheet_label: str, n_tickets: int, n_cols: int, title_row: int = 1) -> None:
    """Merged colored band: 'BB LOS  •  14 tickets  •  as of 28 Apr 2026'."""
    today = datetime.now().strftime("%d %b %Y")
    text = f"{sheet_label}   \u2022   {n_tickets} ticket{'s' if n_tickets != 1 else ''}   \u2022   as of {today}"
    end_col = max(n_cols, 1)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=end_col)
    cell = ws.cell(row=title_row, column=1)
    cell.value = text
    cell.fill = PatternFill(start_color=TITLE_BAR_FILL_HEX, end_color=TITLE_BAR_FILL_HEX, fill_type="solid")
    cell.font = Font(bold=True, color=TITLE_BAR_FONT_HEX, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[title_row].height = 26


def _shade_rows_by_assignee(ws, df: pd.DataFrame, assign_col_name: Optional[str],
                             data_start_row: int, n_cols: int) -> None:
    """Pastel row shading by cleaned assignee name."""
    if df is None or df.empty or assign_col_name is None or assign_col_name not in df.columns:
        return
    pretty = clean_assignee_series(df[assign_col_name])
    unique_names: List[str] = []
    seen: set = set()
    for v in pretty:
        if v not in seen:
            seen.add(v)
            unique_names.append(v)
    color_map = {name: PASTEL_HEX_COLORS[i % len(PASTEL_HEX_COLORS)]
                 for i, name in enumerate(unique_names)}
    for offset, name in enumerate(pretty):
        r = data_start_row + offset
        color = color_map.get(name, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = fill


def _replace_assign_column_with_pretty(ws, df: pd.DataFrame, assign_col_name: Optional[str],
                                        header_row: int, data_start_row: int) -> None:
    """Overwrite the _Assign column cells with cleaned, title-cased names."""
    if df is None or df.empty or assign_col_name is None or assign_col_name not in df.columns:
        return
    target_norm = assign_col_name.strip().lower()
    col_idx = None
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=j).value
        if v is not None and str(v).strip().lower() == target_norm:
            col_idx = j
            break
    if col_idx is None:
        return
    pretty = clean_assignee_series(df[assign_col_name]).tolist()
    for offset, name in enumerate(pretty):
        ws.cell(row=data_start_row + offset, column=col_idx).value = name


def _apply_sla_color_scale(ws, df: pd.DataFrame, header_row: int,
                            data_start_row: int, data_end_row: int,
                            settings: Optional[Dict[str, object]] = None) -> None:
    """5-tier SLA conditional formatting on the SLA column.
    Threshold values come from settings if provided, else from defaults."""
    sla_idx = None
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=j).value
        if v is not None and str(v).strip().lower() == "sla":
            sla_idx = j
            break
    if sla_idx is None or data_end_row < data_start_row:
        return

    s = settings or DEFAULT_SETTINGS
    g = int(s.get("sla_green_max", 1))
    y = int(s.get("sla_yellow_max", 3))
    o = int(s.get("sla_orange_max", 7))
    r_max = int(s.get("sla_red_max", 14))

    col_letter = get_column_letter(sla_idx)
    cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"

    rules = [
        # Highest priority first; stopIfTrue prevents lower rules from firing
        CellIsRule(operator="greaterThan", formula=[str(r_max)], stopIfTrue=True,
                   fill=PatternFill(start_color=SLA_DARKRED_HEX, end_color=SLA_DARKRED_HEX, fill_type="solid"),
                   font=Font(bold=True, color="FFFFFF")),
        CellIsRule(operator="between", formula=[str(o + 1), str(r_max)], stopIfTrue=True,
                   fill=PatternFill(start_color=SLA_RED_HEX, end_color=SLA_RED_HEX, fill_type="solid"),
                   font=Font(bold=True)),
        CellIsRule(operator="between", formula=[str(y + 1), str(o)], stopIfTrue=True,
                   fill=PatternFill(start_color=SLA_ORANGE_HEX, end_color=SLA_ORANGE_HEX, fill_type="solid")),
        CellIsRule(operator="between", formula=[str(g + 1), str(y)], stopIfTrue=True,
                   fill=PatternFill(start_color=SLA_YELLOW_HEX, end_color=SLA_YELLOW_HEX, fill_type="solid")),
        CellIsRule(operator="between", formula=["0", str(g)], stopIfTrue=True,
                   fill=PatternFill(start_color=SLA_GREEN_HEX, end_color=SLA_GREEN_HEX, fill_type="solid")),
    ]
    for rule in rules:
        ws.conditional_formatting.add(cell_range, rule)

    for r in range(data_start_row, data_end_row + 1):
        ws.cell(row=r, column=sla_idx).alignment = Alignment(horizontal="center", vertical="center")


def _write_assign_summary_pretty(ws, df: pd.DataFrame, assign_col_name: Optional[str],
                                  header_row: int, n_data_cols: int) -> None:
    """Mini Assignee/Count table at the top-right, with cleaned names."""
    if df is None or df.empty:
        return
    if assign_col_name is not None and assign_col_name in df.columns:
        labels = clean_assignee_series(df[assign_col_name])
    else:
        labels = pd.Series(["(Unassigned)"] * len(df))

    counts = (
        pd.DataFrame({"_Assign": labels})
        .groupby("_Assign", dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(["Count", "_Assign"], ascending=[False, True])
    )

    start_col = n_data_cols + 2
    border = _all_borders()
    header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FONT_HEX, size=11)

    r = header_row
    h1 = ws.cell(row=r, column=start_col)
    h1.value = "Assignee"
    h1.fill = header_fill
    h1.font = header_font
    h1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h1.border = border

    h2 = ws.cell(row=r, column=start_col + 1)
    h2.value = "Count"
    h2.fill = header_fill
    h2.font = header_font
    h2.alignment = Alignment(horizontal="center", vertical="center")
    h2.border = border

    r += 1
    for _, row in counts.iterrows():
        a = ws.cell(row=r, column=start_col)
        a.value = str(row["_Assign"])
        a.alignment = Alignment(horizontal="left", indent=1)
        a.border = border
        c = ws.cell(row=r, column=start_col + 1)
        c.value = int(row["Count"])
        c.alignment = Alignment(horizontal="center")
        c.border = border
        r += 1

    r += 1
    t1 = ws.cell(row=r, column=start_col)
    t1.value = "TOTAL"
    t1.font = Font(bold=True, color=HEADER_FONT_HEX)
    t1.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t1.alignment = Alignment(horizontal="left", indent=1)
    t1.border = border
    t2 = ws.cell(row=r, column=start_col + 1)
    t2.value = int(counts["Count"].sum()) if not counts.empty else 0
    t2.font = Font(bold=True, color=HEADER_FONT_HEX)
    t2.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t2.alignment = Alignment(horizontal="center")
    t2.border = border

    ws.column_dimensions[get_column_letter(start_col)].width = 26
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 8


def apply_sheet_styling(ws, df: pd.DataFrame, assign_col_name: Optional[str],
                        sheet_label: str,
                        settings: Optional[Dict[str, object]] = None) -> None:
    """Full visual treatment for a single data sheet.

    AFTER pandas writes the dataframe (header at row 1, data from row 2),
    we INSERT a row at top to make the title bar at row 1, headers at row 2,
    data from row 3.
    """
    if df is None or df.empty:
        return
    n_data_cols = len(df.columns)

    ws.insert_rows(1)

    header_row = 2
    data_start_row = 3
    data_end_row = data_start_row + len(df) - 1

    _apply_title_bar(ws, sheet_label, n_tickets=len(df), n_cols=n_data_cols, title_row=1)
    _apply_header_styling(ws, header_row=header_row, n_cols=n_data_cols)
    _replace_assign_column_with_pretty(ws, df, assign_col_name, header_row, data_start_row)
    _shade_rows_by_assignee(ws, df, assign_col_name, data_start_row, n_data_cols)

    border = _all_borders()
    for r in range(header_row, data_end_row + 1):
        for c in range(1, n_data_cols + 1):
            ws.cell(row=r, column=c).border = border

    _apply_sla_color_scale(ws, df, header_row, data_start_row, data_end_row, settings=settings)
    _autosize_columns(ws, df)
    ws.freeze_panes = ws.cell(row=data_start_row, column=1).coordinate

    last_col_letter = get_column_letter(n_data_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end_row}"

    color = tab_color_for(sheet_label)
    if color:
        ws.sheet_properties.tabColor = color

    _write_assign_summary_pretty(ws, df, assign_col_name, header_row=header_row,
                                  n_data_cols=n_data_cols)


# ============================================================================
# SECTION 6   Adaptive Summary layout (brick-wall packing)
# ============================================================================
#
# The classic 4-up grid wastes space when blocks have wildly different heights
# (e.g. Power has 5+ assignees, BB on CRC Errors has 1). This layout treats
# the sheet as a fixed-column grid and packs blocks into the columns greedily,
# placing each new block at the column with the LOWEST current bottom row.
# That eliminates blank rows under short blocks.
#
# Block layout per column slot (2 worksheet columns + 1 spacer):
#     row N      Title   (colored, merged across both data cols)
#     row N+1    Sub-header (Assignee | Count)
#     row N+2..  Data rows (one per assignee)
#     row K      Total row (KPI accent)
#
# After packing, the Grand Total panel gets placed in whichever region has
# the most leftover space — to the right if rows were tall, below if wide.

def _block_height(df_block: pd.DataFrame, assign_col_name: Optional[str]) -> int:
    """Predict how many rows a block will occupy when written."""
    if df_block is None or df_block.empty:
        return 0
    counts = _block_counts(df_block, assign_col_name)
    # title (1) + subheader (1) + data rows + total (1) = 3 + rows
    # plus SLA breakdown: header (1) + 7 brackets = 8 extra rows
    return 3 + len(counts) + 8


def _sla_breakdown_counts(df_block: pd.DataFrame) -> List[Tuple[str, int]]:
    """Compute SLA bracket counts for a dataframe bucket.
    Brackets: <24hrs, 1 day, 2 days, 3 days, 4 days, 5 days, >5 days.
    Uses the 'creation' column if present, otherwise the 'SLA' column.
    Returns list of (label, count) tuples.
    """
    brackets = [
        ("<24 Hrs", lambda d: d < 1),
        ("1 Day",   lambda d: d == 1),
        ("2 Days",  lambda d: d == 2),
        ("3 Days",  lambda d: d == 3),
        ("4 Days",  lambda d: d == 4),
        ("5 Days",  lambda d: d == 5),
        (">5 Days", lambda d: d > 5),
    ]

    # Try to compute days from 'creation' first, then fall back to 'SLA' column
    sla_days: Optional[pd.Series] = None
    creation_col = _find_column_case_insensitive(df_block, CREATION_COLUMN_NORMALIZED)
    if creation_col is not None:
        try:
            creation_dt = pd.to_datetime(df_block[creation_col], errors="coerce")
            now = pd.Timestamp.now()
            sla_days = (now - creation_dt).dt.days
        except Exception:
            sla_days = None

    if sla_days is None:
        sla_col = None
        for col in df_block.columns:
            if str(col).strip().lower() == "sla":
                sla_col = col
                break
        if sla_col is not None:
            try:
                sla_days = pd.to_numeric(df_block[sla_col], errors="coerce")
            except Exception:
                sla_days = None

    if sla_days is None or sla_days.isna().all():
        return [(label, 0) for label, _ in brackets]

    result = []
    for label, fn in brackets:
        count = int((sla_days.dropna().apply(fn)).sum())
        result.append((label, count))
    return result


def _write_sla_breakdown(ws, df_block: pd.DataFrame, start_row: int, start_col: int) -> int:
    """Write SLA breakdown rows below a summary block. Returns the last row used."""
    sla_counts = _sla_breakdown_counts(df_block)
    border = _all_borders()

    SLA_BRACKET_COLORS = {
        "<24 Hrs": SLA_GREEN_HEX,
        "1 Day":   SLA_GREEN_HEX,
        "2 Days":  SLA_YELLOW_HEX,
        "3 Days":  SLA_YELLOW_HEX,
        "4 Days":  SLA_ORANGE_HEX,
        "5 Days":  SLA_ORANGE_HEX,
        ">5 Days": SLA_RED_HEX,
    }

    # SLA Breakdown header row
    r = start_row
    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col + 1)
    hdr = ws.cell(row=r, column=start_col)
    hdr.value = "SLA Breakdown"
    hdr.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    hdr.font = Font(bold=True, color="FFFFFF", size=9)
    hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    hdr.border = border
    ws.cell(row=r, column=start_col + 1).border = border
    ws.row_dimensions[r].height = 14
    r += 1

    for label, count in sla_counts:
        color = SLA_BRACKET_COLORS.get(label, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        a = ws.cell(row=r, column=start_col)
        a.value = label
        a.fill = fill
        a.font = Font(size=9)
        a.alignment = Alignment(horizontal="left", indent=2)
        a.border = border
        c = ws.cell(row=r, column=start_col + 1)
        c.value = count
        c.fill = fill
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        ws.row_dimensions[r].height = 13
        r += 1

    return r - 1


def _write_summary_block(ws, title: str, df_block: pd.DataFrame,
                          assign_col_name: Optional[str],
                          start_row: int, start_col: int) -> int:
    """Write a single mini-table for one bucket. Returns the last row used."""
    if df_block is None or df_block.empty:
        return start_row - 1

    counts = _block_counts(df_block, assign_col_name)
    border = _all_borders()
    tab_color = tab_color_for(title)

    # Title row (colored band, spans both columns)
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col + 1)
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.fill = PatternFill(start_color=tab_color, end_color=tab_color, fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title_cell.border = border
    ws.cell(row=start_row, column=start_col + 1).border = border
    ws.row_dimensions[start_row].height = 18

    # Sub-header
    r = start_row + 1
    sh1 = ws.cell(row=r, column=start_col)
    sh1.value = "Assignee"
    sh1.fill = PatternFill(start_color=SUB_HEADER_FILL_HEX, end_color=SUB_HEADER_FILL_HEX, fill_type="solid")
    sh1.font = Font(bold=True, size=10)
    sh1.alignment = Alignment(horizontal="left", indent=1)
    sh1.border = border

    sh2 = ws.cell(row=r, column=start_col + 1)
    sh2.value = "Count"
    sh2.fill = PatternFill(start_color=SUB_HEADER_FILL_HEX, end_color=SUB_HEADER_FILL_HEX, fill_type="solid")
    sh2.font = Font(bold=True, size=10)
    sh2.alignment = Alignment(horizontal="center")
    sh2.border = border

    r += 1
    for _, crow in counts.iterrows():
        a = ws.cell(row=r, column=start_col)
        a.value = str(crow["_Assign"]) if pd.notna(crow["_Assign"]) else ""
        a.alignment = Alignment(horizontal="left", indent=1)
        a.border = border
        c = ws.cell(row=r, column=start_col + 1)
        c.value = int(crow["Count"]) if pd.notna(crow["Count"]) else 0
        c.alignment = Alignment(horizontal="center")
        c.border = border
        r += 1

    # Total row
    t1 = ws.cell(row=r, column=start_col)
    t1.value = "Total"
    t1.font = Font(bold=True, color="FFFFFF")
    t1.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t1.alignment = Alignment(horizontal="left", indent=1)
    t1.border = border
    t2 = ws.cell(row=r, column=start_col + 1)
    t2.value = int(len(df_block))
    t2.font = Font(bold=True, color="FFFFFF")
    t2.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t2.alignment = Alignment(horizontal="center")
    t2.border = border
    r += 1

    # SLA Breakdown section below the assignee/count block
    r = _write_sla_breakdown(ws, df_block, start_row=r, start_col=start_col)

    return r


def pack_blocks_adaptive(ws, blocks: List[Tuple[str, pd.DataFrame]],
                          assign_col_name: Optional[str],
                          start_row: int,
                          start_col: int,
                          n_columns: int,
                          col_step: int = 3,
                          row_gap: int = 1) -> Tuple[int, int]:
    """Greedy brick-wall packer.

    Places each block in the column with the lowest current bottom row,
    so short blocks fill in under the previous tall ones.

    Returns (max_row_used, max_col_used).
    """
    # Track the next available row in each of the n_columns column slots
    col_bottom = [start_row] * n_columns
    max_row_used = start_row
    max_col_used = start_col

    # Skip empty blocks early so we don't claim a slot for them
    populated_blocks = [(t, d) for t, d in blocks if d is not None and not d.empty]

    for title, df_block in populated_blocks:
        # Pick column with smallest current bottom (ties -> leftmost)
        slot_idx = min(range(n_columns), key=lambda i: (col_bottom[i], i))
        place_row = col_bottom[slot_idx]
        place_col = start_col + slot_idx * col_step

        last_row = _write_summary_block(ws, title, df_block, assign_col_name,
                                         place_row, place_col)
        new_bottom = last_row + 1 + row_gap
        col_bottom[slot_idx] = new_bottom

        if last_row > max_row_used:
            max_row_used = last_row
        if place_col + 1 > max_col_used:
            max_col_used = place_col + 1

    return max_row_used, max_col_used, col_bottom


def write_grand_total_panel(ws, all_buckets: List[Tuple[str, pd.DataFrame]],
                             assign_col_name: Optional[str],
                             start_row: int, start_col: int,
                             title: str = "GRAND TOTAL — Tickets per Assignee (all buckets)",
                             title_span: int = 2) -> int:
    """Bottom/right-side panel: tickets per assignee aggregated across all buckets."""
    if not all_buckets:
        return start_row
    big = pd.concat([df for _, df in all_buckets if df is not None and not df.empty],
                    ignore_index=True)
    if big.empty:
        return start_row

    if assign_col_name is not None and assign_col_name in big.columns:
        labels = clean_assignee_series(big[assign_col_name])
    else:
        labels = pd.Series(["(Unassigned)"] * len(big))

    counts = (
        pd.DataFrame({"_Assign": labels})
        .groupby("_Assign", dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(["Count", "_Assign"], ascending=[False, True])
    )

    border = _all_borders()

    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col + title_span - 1)
    t = ws.cell(row=start_row, column=start_col)
    t.value = title
    t.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t.border = border
    ws.row_dimensions[start_row].height = 22

    r = start_row + 1
    sh1 = ws.cell(row=r, column=start_col)
    sh1.value = "Assignee"
    sh1.fill = PatternFill(start_color=SUB_HEADER_FILL_HEX, end_color=SUB_HEADER_FILL_HEX, fill_type="solid")
    sh1.font = Font(bold=True)
    sh1.alignment = Alignment(horizontal="left", indent=1)
    sh1.border = border

    sh2 = ws.cell(row=r, column=start_col + 1)
    sh2.value = "Count"
    sh2.fill = PatternFill(start_color=SUB_HEADER_FILL_HEX, end_color=SUB_HEADER_FILL_HEX, fill_type="solid")
    sh2.font = Font(bold=True)
    sh2.alignment = Alignment(horizontal="center")
    sh2.border = border

    r += 1
    grand_total = 0
    for _, crow in counts.iterrows():
        a = ws.cell(row=r, column=start_col)
        a.value = str(crow["_Assign"])
        a.alignment = Alignment(horizontal="left", indent=1)
        a.border = border
        c = ws.cell(row=r, column=start_col + 1)
        c.value = int(crow["Count"])
        c.alignment = Alignment(horizontal="center")
        c.border = border
        grand_total += int(crow["Count"])
        r += 1

    t1 = ws.cell(row=r, column=start_col)
    t1.value = "TOTAL"
    t1.font = Font(bold=True, color="FFFFFF")
    t1.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t1.alignment = Alignment(horizontal="left", indent=1)
    t1.border = border
    t2 = ws.cell(row=r, column=start_col + 1)
    t2.value = grand_total
    t2.font = Font(bold=True, color="FFFFFF")
    t2.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    t2.alignment = Alignment(horizontal="center")
    t2.border = border

    return r


def write_kpi_strip(ws, all_buckets: List[Tuple[str, pd.DataFrame]],
                    assign_col_name: Optional[str], start_row: int,
                    n_kpi_cards: int = 5,
                    extra_kpis: Optional[List[Tuple[str, str]]] = None) -> int:
    """Top KPI strip: 5 cards in a row.
    Returns the next free row after the strip."""
    if all_buckets:
        big = pd.concat([df for _, df in all_buckets if df is not None and not df.empty],
                        ignore_index=True)
    else:
        big = pd.DataFrame()

    total_tickets = len(big)

    open_count = 0
    for col in big.columns:
        if str(col).strip().lower() == "status":
            open_count = int((big[col].astype(str).str.strip().str.lower() == "open").sum())
            break

    top_assignee = "—"
    if assign_col_name is not None and assign_col_name in big.columns and not big.empty:
        pretty = clean_assignee_series(big[assign_col_name])
        non_unassigned = pretty[pretty != "(Unassigned)"]
        if not non_unassigned.empty:
            top_assignee = non_unassigned.value_counts().index[0]
        elif not pretty.empty:
            top_assignee = pretty.value_counts().index[0]

    avg_sla_str = "—"
    oldest_sla_str = "—"
    creation_col = None
    for col in big.columns:
        if str(col).strip().lower() == "creation":
            creation_col = col
            break
    if creation_col is not None and not big.empty:
        try:
            dt = pd.to_datetime(big[creation_col], errors="coerce")
            now = pd.Timestamp.now()
            days = (now - dt).dt.days.dropna()
            if len(days) > 0:
                avg_sla_str = f"{days.mean():.1f} days"
                oldest_sla_str = f"{int(days.max())} days"
        except Exception:
            pass

    kpis: List[Tuple[str, str]] = [
        ("Total Tickets", f"{total_tickets}"),
        ("Open", f"{open_count}"),
        ("Top Assignee", top_assignee),
        ("Avg SLA Age", avg_sla_str),
        ("Oldest Ticket", oldest_sla_str),
    ]
    if extra_kpis:
        kpis = kpis + list(extra_kpis)

    label_fill = PatternFill(start_color=KPI_FILL_HEX, end_color=KPI_FILL_HEX, fill_type="solid")
    label_font = Font(bold=True, color="595959", size=10)
    value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    value_font = Font(bold=True, color=KPI_ACCENT_HEX, size=16)
    border = _all_borders()

    col_step = 3
    for idx, (label, value) in enumerate(kpis):
        c0 = 1 + idx * col_step

        ws.merge_cells(start_row=start_row, start_column=c0,
                       end_row=start_row, end_column=c0 + 1)
        lbl = ws.cell(row=start_row, column=c0)
        lbl.value = label.upper()
        lbl.fill = label_fill
        lbl.font = label_font
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border = border

        ws.merge_cells(start_row=start_row + 1, start_column=c0,
                       end_row=start_row + 1, end_column=c0 + 1)
        val = ws.cell(row=start_row + 1, column=c0)
        val.value = value
        val.fill = value_fill
        val.font = value_font
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = border

    ws.row_dimensions[start_row].height = 18
    ws.row_dimensions[start_row + 1].height = 36

    return start_row + 3


def apply_summary_styling_adaptive(
    ws,
    all_buckets: List[Tuple[str, pd.DataFrame]],
    assign_col_name: Optional[str],
    input_path: Path,
    report_kind: str,
    n_columns: int = 4,
) -> None:
    """Build a Summary sheet using adaptive packing.

    1. Header band (title + source line)
    2. KPI strip (5 cards)
    3. Adaptive bucket grid — short blocks fill gaps under tall ones
    4. Grand Total panel — placed to the RIGHT of the bucket grid if there's
       horizontal room, otherwise BELOW it
    """
    today = datetime.now().strftime("%d %B %Y")
    title_text = f"{report_kind.upper()} REPORT  —  {today}"

    # Compute total width we'll use for KPIs and grid.
    # KPI strip: 5 cards * 3 cols = 15 cols.  We fit the bucket grid inside that.
    total_width = 1 + (5 * 3) - 1  # cols 1..15

    # Title row (row 1) — the only header. Compact, no source line, no KPI strip.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_width)
    title = ws.cell(row=1, column=1)
    title.value = title_text
    title.fill = PatternFill(start_color=KPI_ACCENT_HEX, end_color=KPI_ACCENT_HEX, fill_type="solid")
    title.font = Font(bold=True, color="FFFFFF", size=18)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Adaptive packing of bucket blocks — starts at row 3 directly under the title
    grid_start_row = 3
    grid_start_col = 1
    col_step = 3

    populated_blocks = [(t, d) for t, d in all_buckets if d is not None and not d.empty]

    if populated_blocks:
        max_row_used, max_col_used, col_bottom = pack_blocks_adaptive(
            ws, populated_blocks, assign_col_name,
            start_row=grid_start_row,
            start_col=grid_start_col,
            n_columns=n_columns,
            col_step=col_step,
            row_gap=1,
        )
    else:
        max_row_used = grid_start_row
        max_col_used = grid_start_col
        col_bottom = [grid_start_row] * n_columns

    # Grand Total placement: prefer placing it as a 5th column on the right of
    # the bucket grid. The summary KPI strip already reserves 5 column slots
    # (cols 1-15). Using slot 5 (col 13) for the Grand Total mirrors the
    # reference design and avoids dumping it below.
    #
    # We only fall back to "below the grid" if the GT panel is taller than
    # any reasonable side placement.

    gt_blocks = populated_blocks
    if gt_blocks:
        big_combined = pd.concat([df for _, df in gt_blocks if df is not None and not df.empty],
                                  ignore_index=True)
        if assign_col_name and assign_col_name in big_combined.columns:
            n_assignees = big_combined[assign_col_name].astype(str).map(clean_assignee).nunique()
        else:
            n_assignees = 1
    else:
        n_assignees = 0
    gt_height = 3 + n_assignees  # title + sub-header + rows + total

    # Slot 5 is the dedicated right-side column for Grand Total
    side_col = grid_start_col + n_columns * col_step  # cols 1,4,7,10 → 13
    side_start_row = grid_start_row

    # Use side placement when GT height does not blow up the layout drastically.
    # In practice, GT should always fit on the side because the KPI strip
    # already reserves 5 column slots horizontally.
    write_grand_total_panel(ws, gt_blocks, assign_col_name,
                             start_row=side_start_row,
                             start_col=side_col,
                             title_span=2)

    # Column widths
    for c in range(1, total_width + 1):
        letter = get_column_letter(c)
        if c in (1, 4, 7, 10, 13):
            ws.column_dimensions[letter].width = 26
        elif c in (2, 5, 8, 11, 14):
            ws.column_dimensions[letter].width = 9
        else:
            ws.column_dimensions[letter].width = 2

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = TAB_COLORS.get("Summary", "FFD700")


# ============================================================================
# SECTION 7   ENTERPRISE pipeline
# ============================================================================

ENTERPRISE_BUCKET_ORDER: List[str] = [
    # Order rows are CLASSIFIED in (priority, first-match wins)
    "Equipment TTs",
    "Power",
    "BB LOS",
    "BB on Extreme Low RX",
    "BB on Low RX",
    "BB on CRC Errors",
    "6150|6120|1050-LOS",
    "6150|6120|1050-Extreme Low RX",
    "6150|6120|1050-Low RX",
    "6150|6120|1050-CRC",
    "Access Ring-LOS",
    "Access Ring-Low RX",
    "Access Ring-Offline",
]


def _classify_enterprise_buckets(
    all_df: pd.DataFrame,
    s_subj: pd.Series,
    s_cat1: pd.Series,
    s_cat3: pd.Series,
    s_cat4: pd.Series,
    s_desc: pd.Series,
    s_any: pd.Series,
) -> Dict[str, pd.DataFrame]:
    """Bucket rows in priority order. First match wins (no double-counting)."""
    if all_df is None or all_df.empty:
        return {name: pd.DataFrame() for name in ENTERPRISE_BUCKET_ORDER}

    zero_mask = pd.Series(False, index=all_df.index)

    def cat1_has_all(*words: str) -> pd.Series:
        m = pd.Series(True, index=s_cat1.index)
        for w in words:
            m = m & s_cat1.str.contains(w, na=False)
        return m

    def cat1_has_any(*words: str) -> pd.Series:
        m = pd.Series(False, index=s_cat1.index)
        for w in words:
            m = m | s_cat1.str.contains(w, na=False)
        return m

    def subj_starts(prefix: str) -> pd.Series:
        return s_subj.str.startswith(prefix.lower(), na=False)

    def subj_contains_any(words: List[str]) -> pd.Series:
        m = pd.Series(False, index=s_subj.index)
        for w in words:
            m = m | s_subj.str.contains(w.lower(), na=False)
        return m

    def s_contains(series: pd.Series, pattern: str, regex: bool = True) -> pd.Series:
        if len(series) == 0:
            return zero_mask
        return series.str.contains(pattern, regex=regex, na=False)

    m_equipment = cat1_has_any("equipment") | s_contains(s_any, r"\bequipment\b")
    m_power = cat1_has_any("power")

    m_bb = (
        subj_starts("bb")
        | s_contains(s_cat3, r"(?:^|\b)(?:ip\backbone|bb)(?:\b|$)")
        | s_contains(s_any, r"(?:^|\b)(?:backbone|bb)(?:\b|$)")
    )
    m_sdh = (
        subj_starts("sdh:")
        | s_contains(s_cat3, r"(?:^|\b)(?:sdh|access\s*ring)(?:\b|$)")
        | s_contains(s_any, r"(?:^|\b)(?:sdh|access\s*ring)(?:\b|$)")
    )

    ran_pat = r"(?:ran/\s*)?(?:6150|6120|1050)"
    m_ran = (
        subj_contains_any(["ran/6150", "ran/6120", "ran/1050", "6150", "6120", "1050"])
        | s_contains(s_desc, ran_pat)
        | s_contains(s_cat4, ran_pat)
        | s_contains(s_any, ran_pat)
    )

    m_los = cat1_has_any("los") | cat1_has_all("fiber", "cut") | s_contains(s_any, r"\blos\b|fiber\s*cut")
    m_extlow = (
        cat1_has_all("extreme", "low", "rx")
        | cat1_has_any("extreme low rx")
        | s_contains(s_any, r"extreme\s*low\s*rx")
    )
    m_lowrx = (
        (cat1_has_all("low", "rx") | cat1_has_any("lowrx") | s_contains(s_any, r"\blow\s*rx\b|\blowrx\b"))
        & (~cat1_has_any("extreme") & ~s_contains(s_any, r"extreme"))
    )
    m_crc = cat1_has_any("crc") | s_contains(s_any, r"\bcrc\b")
    m_offline = cat1_has_any("offline", "isolation") | s_contains(s_any, r"\boffline\b|\bisolation\b")

    buckets: List[Tuple[str, pd.Series]] = [
        ("Equipment TTs", m_equipment),
        ("Power", m_power),
        ("BB LOS", m_bb & m_los),
        ("BB on Extreme Low RX", m_bb & m_extlow),
        ("BB on Low RX", m_bb & m_lowrx),
        ("BB on CRC Errors", m_bb & m_crc),
        ("6150|6120|1050-LOS", m_ran & m_los),
        ("6150|6120|1050-Extreme Low RX", m_ran & m_extlow),
        ("6150|6120|1050-Low RX", m_ran & m_lowrx),
        ("6150|6120|1050-CRC", m_ran & m_crc),
        ("Access Ring-LOS", m_sdh & m_los),
        ("Access Ring-Low RX", m_sdh & m_lowrx),
        ("Access Ring-Offline", m_sdh & m_offline),
    ]

    assigned = pd.Series(False, index=all_df.index)
    bucket_frames: Dict[str, pd.DataFrame] = {}
    for name, m in buckets:
        mask = m & (~assigned)
        if mask.any():
            bucket_frames[name] = all_df.loc[mask].reset_index(drop=True)
            assigned |= mask
        else:
            bucket_frames[name] = pd.DataFrame()
    return bucket_frames


def enterprise_preview(input_path: Path) -> Dict[str, object]:
    """Quick read-only preview of what the Enterprise pipeline would produce."""
    sheets = pd.read_excel(input_path, sheet_name=None)
    total_rows = sum(len(df) for df in sheets.values())

    cleaned_frames: List[pd.DataFrame] = []
    closed_total = 0
    for _, df in sheets.items():
        df2, _ = _drop_columns(df)
        df3, closed_n, _ = _filter_closed_status(df2)
        closed_total += closed_n
        cleaned_frames.append(df3)

    all_df = pd.concat(cleaned_frames, ignore_index=True) if cleaned_frames else pd.DataFrame()

    s_any_full = pd.Series([""] * len(all_df), index=all_df.index, dtype=str) if not all_df.empty else pd.Series([], dtype=str)
    for col in list(all_df.columns):
        s_any_full = (s_any_full + " " + _norm_series(all_df, col)).str.strip()

    gpon_dropped = 0
    if len(s_any_full) > 0:
        m_gpon = s_any_full.str.contains(r"\bgpon\b", regex=True, na=False)
        gpon_dropped = int(m_gpon.sum())
        all_df = all_df.loc[~m_gpon].reset_index(drop=True)

    subj_col = _find_column_case_insensitive(all_df, SUBJECT_COLUMN_NORMALIZED)
    cat1_col = _find_column_case_insensitive(all_df, CATEGORY1_COLUMN_NORMALIZED)
    cat3_col = _find_column_case_insensitive(all_df, CATEGORY3_COLUMN_NORMALIZED)
    cat4_col = _find_column_case_insensitive(all_df, CATEGORY4_COLUMN_NORMALIZED)
    desc_col = _find_column_case_insensitive(all_df, DESCRIPTION_COLUMN_NORMALIZED)

    s_subj = _norm_series(all_df, subj_col)
    s_cat1 = _norm_series(all_df, cat1_col)
    s_cat3 = _norm_series(all_df, cat3_col)
    s_cat4 = _norm_series(all_df, cat4_col)
    s_desc = _norm_series(all_df, desc_col)
    s_any = pd.Series([""] * len(all_df), index=all_df.index, dtype=str) if not all_df.empty else pd.Series([], dtype=str)
    for col in list(all_df.columns):
        s_any = (s_any + " " + _norm_series(all_df, col)).str.strip()

    bucket_frames = _classify_enterprise_buckets(all_df, s_subj, s_cat1, s_cat3, s_cat4, s_desc, s_any)
    bucket_counts = {name: len(df) for name, df in bucket_frames.items()}

    cesr_count = 0
    if len(s_any) > 0:
        cesr_count = int(s_any.str.contains(r"\bcesr\b", regex=True, na=False).sum())

    cleaned_total = sum(len(df) for df in cleaned_frames)

    return {
        "sheet_names": list(sheets.keys()),
        "total_rows": total_rows,
        "kept_rows": len(all_df),
        "dropped_closed": closed_total,
        "dropped_gpon": gpon_dropped,
        "bucket_counts": bucket_counts,
        "cesr_count": cesr_count,
        "columns": list(cleaned_frames[0].columns) if cleaned_frames else [],
        "cleaned_total": cleaned_total,
    }


def enterprise_process(input_path: Path, output_path: Path,
                        progress_callback: Optional[Callable] = None,
                        settings: Optional[Dict[str, object]] = None) -> Dict[str, dict]:
    """Generate the Enterprise report. progress_callback(phase, current, total)."""

    def _progress(phase: str, current: int = 0, total: int = 1) -> None:
        if progress_callback is not None:
            try:
                progress_callback(phase, current, total)
            except Exception:
                pass

    _progress("load", 0, 1)
    sheets = pd.read_excel(input_path, sheet_name=None)
    _progress("load", 1, 1)

    cleaned_frames: List[pd.DataFrame] = []
    stats: Dict[str, dict] = {}
    n_sheets = max(1, len(sheets))
    for i, (sheet_name, df) in enumerate(sheets.items()):
        orig_rows = len(df)
        df2, dropped_cols = _drop_columns(df)
        df3, closed_removed, status_col = _filter_closed_status(df2)
        cleaned_frames.append(df3)
        stats[sheet_name] = {
            "original_rows": orig_rows,
            "final_rows": len(df3),
            "dropped_columns": dropped_cols,
            "status_column": status_col,
            "closed_rows_removed": closed_removed,
        }
        _progress("filter", i + 1, n_sheets)

    all_df = pd.concat(cleaned_frames, ignore_index=True) if cleaned_frames else pd.DataFrame()

    subj_col = _find_column_case_insensitive(all_df, SUBJECT_COLUMN_NORMALIZED)
    cat1_col = _find_column_case_insensitive(all_df, CATEGORY1_COLUMN_NORMALIZED)
    assign_col = _find_column_case_insensitive(all_df, ASSIGN_COLUMN_NORMALIZED)
    cat3_col = _find_column_case_insensitive(all_df, CATEGORY3_COLUMN_NORMALIZED)
    cat4_col = _find_column_case_insensitive(all_df, CATEGORY4_COLUMN_NORMALIZED)
    desc_col = _find_column_case_insensitive(all_df, DESCRIPTION_COLUMN_NORMALIZED)

    s_subj = _norm_series(all_df, subj_col)
    s_cat1 = _norm_series(all_df, cat1_col)
    s_cat3 = _norm_series(all_df, cat3_col)
    s_cat4 = _norm_series(all_df, cat4_col)
    s_desc = _norm_series(all_df, desc_col)

    s_any = pd.Series([""] * len(all_df), index=all_df.index, dtype=str) if not all_df.empty else pd.Series([], dtype=str)
    for col in list(all_df.columns):
        s_any = (s_any + " " + _norm_series(all_df, col)).str.strip()

    # GPON exclusion
    if len(s_any) > 0:
        m_excl_gpon = s_any.str.contains(r"\bgpon\b", regex=True, na=False)
        if m_excl_gpon.any():
            all_df = all_df.loc[~m_excl_gpon].reset_index(drop=True)
            s_subj = _norm_series(all_df, subj_col)
            s_cat1 = _norm_series(all_df, cat1_col)
            s_cat3 = _norm_series(all_df, cat3_col)
            s_cat4 = _norm_series(all_df, cat4_col)
            s_desc = _norm_series(all_df, desc_col)
            s_any = pd.Series([""] * len(all_df), index=all_df.index, dtype=str) if not all_df.empty else pd.Series([], dtype=str)
            for col in list(all_df.columns):
                s_any = (s_any + " " + _norm_series(all_df, col)).str.strip()

    _progress("classify", 0, 1)
    bucket_frames = _classify_enterprise_buckets(all_df, s_subj, s_cat1, s_cat3, s_cat4, s_desc, s_any)

    if (all_df is not None and not all_df.empty) and all(fr.empty for fr in bucket_frames.values()):
        bucket_frames["All Tickets"] = all_df.reset_index(drop=True)

    cesr_df = pd.DataFrame()
    if not all_df.empty and len(s_any) > 0:
        m_cesr = s_any.str.contains(r"\bcesr\b", regex=True, na=False)
        cesr_df = all_df.loc[m_cesr].reset_index(drop=True)
    _progress("classify", 1, 1)

    _progress("write", 0, 1)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used: set = set()

        # Write fixed-order buckets
        non_empty_buckets: List[Tuple[str, pd.DataFrame]] = []
        for name in ENTERPRISE_BUCKET_ORDER:
            dfb = bucket_frames.get(name, pd.DataFrame())
            if dfb is None or dfb.empty:
                continue
            df_out = _append_sla_days(dfb)
            df_out.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            apply_sheet_styling(ws, df_out, assign_col_name=assign_col, sheet_label=name, settings=settings)
            used.add(name)
            non_empty_buckets.append((name, dfb))

        # Fallback / extras
        for name, dfb in bucket_frames.items():
            if name in used or dfb is None or dfb.empty:
                continue
            df_out = _append_sla_days(dfb)
            df_out.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            apply_sheet_styling(ws, df_out, assign_col_name=assign_col, sheet_label=name, settings=settings)
            used.add(name)
            non_empty_buckets.append((name, dfb))

        # CESR sheet
        if not cesr_df.empty:
            df_out = _append_sla_days(cesr_df)
            df_out.to_excel(writer, sheet_name="CESR", index=False)
            ws = writer.sheets["CESR"]
            apply_sheet_styling(ws, df_out, assign_col_name=assign_col, sheet_label="CESR", settings=settings)
            non_empty_buckets.append(("CESR", cesr_df))

        # Reorder for Summary by priority (most important first)
        priority_index = {name: i for i, name in enumerate(ENTERPRISE_BUCKET_PRIORITY)}
        non_empty_buckets_sorted = sorted(
            non_empty_buckets,
            key=lambda kv: (priority_index.get(kv[0], 999), -len(kv[1])),
        )

        # Summary
        sum_name = "Summary"
        pd.DataFrame().to_excel(writer, sheet_name=sum_name, index=False)
        ws_sum = writer.sheets[sum_name]
        apply_summary_styling_adaptive(
            ws_sum,
            non_empty_buckets_sorted,
            assign_col_name=assign_col,
            input_path=input_path,
            report_kind="Enterprise",
            n_columns=4,
        )

        # Move Summary to front
        try:
            wb = writer.book
            if sum_name in writer.sheets:
                ws = writer.sheets[sum_name]
                wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))
        except Exception:
            pass

    _progress("done", 1, 1)
    return stats


# ============================================================================
# SECTION 8   GPON pipeline
# ============================================================================
#
# Sheet structure produced (in order):
#     GPON SUMMARY         - KPI strip + per-bucket adaptive grid + grand total
#     GPON - <Cat1>...     - one sheet per Category 1 of GPON tickets
#     Splitter - <Cat1>... - splitter tickets grouped by Category 1
#     Customer Action      - special status (GPON-only)
#     Under Monitoring     - special status (GPON-only)
#     Temporary Restoration- special status (GPON-only)
#     Double Tickets       - same Subscription+Category 1 (GPON only, with exemptions)
#     GPON LOS - Customer Action
#     MDU - LOS / MDU - Low RX

GPON_BUCKET_PRIORITY: List[str] = [
    "GPON LOS Fiber Cut",
    "GPON Extreme Low RX",
    "GPON Low RX",
    "GPON LOSi",
    "GPON SFi",
    "GPON LOFi",
    "Customer Action",
    "Under Monitoring",
    "Temporary Restoration",
    "Double Tickets",
    "MDU - LOS",
    "MDU - Low RX",
]


def _isolate_gpon_rows(df3: pd.DataFrame) -> Tuple[pd.DataFrame, pd.Series]:
    """Identify which rows are GPON. Replicates original heuristics:
    - Category 3 contains 'GPON'
    - OR Category 3 has a 10-char hex id AND Subject contains a port-like pattern
    - OR Category 3 or Category 2 is empty AND Category 1 contains 'GPON'
    Returns (gpon_subset_df, mask_of_gpon_in_df3).
    """
    if df3 is None or df3.empty:
        return pd.DataFrame(), pd.Series([], dtype=bool)

    cat3_col = _find_column_case_insensitive(df3, CATEGORY3_COLUMN_NORMALIZED)
    cat2_col = _find_column_case_insensitive(df3, CATEGORY2_COLUMN_NORMALIZED)
    cat1_col = _find_column_case_insensitive(df3, CATEGORY1_COLUMN_NORMALIZED)
    subj_col = _find_column_case_insensitive(df3, SUBJECT_COLUMN_NORMALIZED)

    mask = pd.Series(False, index=df3.index)

    # 1) Category 3 contains GPON
    if cat3_col is not None:
        m1 = df3[cat3_col].astype(str).str.contains(GPON_KEYWORD, case=False, na=False)
        mask = mask | m1
    else:
        m1 = pd.Series(False, index=df3.index)

    # 2) Category 3 has 10-char hex AND subject has port-like pattern
    if cat3_col is not None and subj_col is not None:
        cat3_series = df3[cat3_col].astype(str)
        has_hex10 = cat3_series.str.contains(r"\b[0-9A-Fa-f]{10}\b", regex=True, na=False)
        subj_series = df3[subj_col].astype(str)
        port_pattern = r"\b\d+(?:/\d+){1,3}\s*:\s*\d+\b"
        subj_has_port = subj_series.str.contains(port_pattern, regex=True, na=False)
        m2 = has_hex10 & subj_has_port & (~m1)
        mask = mask | m2

    # 3) Cat3 empty OR Cat2 empty, AND Cat1 contains GPON
    if cat1_col is not None:
        c1_has_gpon = df3[cat1_col].astype(str).str.contains(GPON_KEYWORD, case=False, na=False)
        cat3_empty = (df3[cat3_col].isna() | (df3[cat3_col].astype(str).str.strip() == "")) if cat3_col is not None else pd.Series(False, index=df3.index)
        cat2_empty = (df3[cat2_col].isna() | (df3[cat2_col].astype(str).str.strip() == "")) if cat2_col is not None else pd.Series(False, index=df3.index)
        m3 = (cat3_empty | cat2_empty) & c1_has_gpon & (~m1)
        mask = mask | m3

    gpon_subset = df3.loc[mask].reset_index(drop=True)
    return gpon_subset, mask


def _collect_status_subsets(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Pull rows whose Status matches the special-status set.
    Returns {display_name: subset_df}."""
    out: Dict[str, pd.DataFrame] = {disp: pd.DataFrame() for disp in STATUS_SPECIAL_SHEETS.values()}
    if df is None or df.empty:
        return out
    status_col = _find_column_case_insensitive(df, STATUS_COLUMN_NORMALIZED)
    if status_col is None:
        return out
    norm_series = df[status_col].astype(str).str.strip().str.lower()
    for norm, disp in STATUS_SPECIAL_SHEETS.items():
        m = norm_series == norm
        if m.any():
            out[disp] = df.loc[m].reset_index(drop=True)
    return out


def _detect_double_tickets(gpon_all: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Detect rows sharing same (Subscription, Category 1) within GPON data.
    Exempt rows where Customer contains 'jamii telecommunication' or 'jamiil limited'.
    Returns (double_tickets_df, gpon_all_minus_doubles).
    """
    if gpon_all is None or gpon_all.empty:
        return pd.DataFrame(), gpon_all
    sub_col = _find_any_col(gpon_all, SUBSCRIPTION_CANDIDATES)
    cat1_col = _find_column_case_insensitive(gpon_all, CATEGORY1_COLUMN_NORMALIZED)
    cust_col = _find_any_col(gpon_all, CUSTOMER_CANDIDATES)
    if sub_col is None or cat1_col is None:
        return pd.DataFrame(), gpon_all

    sub_norm = gpon_all[sub_col].astype(str).str.strip().str.lower()
    cat_norm = gpon_all[cat1_col].astype(str).str.strip().str.lower()
    has_sub = gpon_all[sub_col].notna() & (gpon_all[sub_col].astype(str).str.strip() != "")
    has_cat = gpon_all[cat1_col].notna() & (gpon_all[cat1_col].astype(str).str.strip() != "")
    base_mask = has_sub & has_cat

    if cust_col is not None:
        cust_norm = gpon_all[cust_col].astype(str).str.strip().str.lower()
        block_cust = (cust_norm.str.contains("jamii telecommunication", na=False)
                      | cust_norm.str.contains("jamiil limited", na=False))
        base_mask &= ~block_cust

    keys_df = pd.DataFrame({"k_sub": sub_norm[base_mask], "k_cat": cat_norm[base_mask]})
    if keys_df.empty:
        return pd.DataFrame(), gpon_all

    counts = keys_df.value_counts(["k_sub", "k_cat"])
    double_keys = {(ks, kc) for (ks, kc), cnt in counts.items() if cnt >= 2}
    if not double_keys:
        return pd.DataFrame(), gpon_all

    mask = pd.Series([(s, c) in double_keys for s, c in zip(sub_norm, cat_norm)],
                    index=gpon_all.index)
    if cust_col is not None:
        cust_norm = gpon_all[cust_col].astype(str).str.strip().str.lower()
        block_cust = (cust_norm.str.contains("jamii telecommunication", na=False)
                      | cust_norm.str.contains("jamiil limited", na=False))
        mask = mask & (~block_cust)

    if not mask.any():
        return pd.DataFrame(), gpon_all

    doubles = gpon_all.loc[mask].reset_index(drop=True)
    remainder = gpon_all.loc[~mask].reset_index(drop=True)
    return doubles, remainder


def _extract_splitter_rows(gpon_all: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Pull rows that contain 'splitter' anywhere out of gpon_all."""
    if gpon_all is None or gpon_all.empty:
        return pd.DataFrame(), gpon_all
    mask = pd.Series(False, index=gpon_all.index)
    for col in gpon_all.columns:
        try:
            m = gpon_all[col].astype(str).str.contains("splitter", case=False, na=False)
        except Exception:
            m = pd.Series(False, index=gpon_all.index)
        mask = mask | m
    if not mask.any():
        return pd.DataFrame(), gpon_all
    splitters = gpon_all.loc[mask].reset_index(drop=True)
    rest = gpon_all.loc[~mask].reset_index(drop=True)
    return splitters, rest


def gpon_preview(input_path: Path) -> Dict[str, object]:
    """Read-only preview of what the GPON pipeline would produce."""
    sheets = pd.read_excel(input_path, sheet_name=None)
    total_rows = sum(len(df) for df in sheets.values())

    cleaned_frames: List[pd.DataFrame] = []
    closed_total = 0
    for _, df in sheets.items():
        df2, _ = _drop_columns(df)
        df3, closed_n, _ = _filter_closed_status(df2)
        closed_total += closed_n
        cleaned_frames.append(df3)
    cleaned_total = sum(len(df) for df in cleaned_frames)

    gpon_frames: List[pd.DataFrame] = []
    for df3 in cleaned_frames:
        gp, _ = _isolate_gpon_rows(df3)
        if not gp.empty:
            gpon_frames.append(gp)
    gpon_all = pd.concat(gpon_frames, ignore_index=True) if gpon_frames else pd.DataFrame()
    gpon_rows = len(gpon_all)

    # Status counts (GPON only)
    status_counts: Dict[str, int] = {}
    for disp in STATUS_SPECIAL_SHEETS.values():
        status_counts[disp] = 0
    for df3 in cleaned_frames:
        if df3.empty:
            continue
        gp, _ = _isolate_gpon_rows(df3)
        if gp.empty:
            continue
        st_subsets = _collect_status_subsets(gp)
        for disp, sd in st_subsets.items():
            status_counts[disp] += len(sd)

    # Detect doubles + splitters in gpon_all
    doubles, gpon_after_doubles = _detect_double_tickets(gpon_all)
    splitters, _ = _extract_splitter_rows(gpon_after_doubles)

    # Per-Category 1 distribution (after subtracting doubles + splitters)
    splitters_after, gpon_main = _extract_splitter_rows(gpon_after_doubles)
    cat1_col = _find_column_case_insensitive(gpon_main, CATEGORY1_COLUMN_NORMALIZED)
    category_counts: Dict[str, int] = {}
    if cat1_col is not None and not gpon_main.empty:
        vc = gpon_main[cat1_col].astype(str).fillna("(Blank)").replace("", "(Blank)").value_counts()
        category_counts = {str(k): int(v) for k, v in vc.items()}

    return {
        "sheet_names": list(sheets.keys()),
        "total_rows": total_rows,
        "cleaned_total": cleaned_total,
        "dropped_closed": closed_total,
        "gpon_rows": gpon_rows,
        "non_gpon_rows": cleaned_total - gpon_rows,
        "status_counts": status_counts,
        "double_tickets": len(doubles),
        "splitter_rows": len(splitters),
        "category_counts": category_counts,
        "columns": list(cleaned_frames[0].columns) if cleaned_frames else [],
    }


def gpon_process(input_path: Path, output_path: Path,
                  progress_callback: Optional[Callable] = None,
                  settings: Optional[Dict[str, object]] = None) -> Dict[str, dict]:
    """Generate the GPON report."""

    def _progress(phase: str, current: int = 0, total: int = 1) -> None:
        if progress_callback is not None:
            try:
                progress_callback(phase, current, total)
            except Exception:
                pass

    _progress("load", 0, 1)
    sheets = pd.read_excel(input_path, sheet_name=None)
    _progress("load", 1, 1)

    cleaned_frames: List[pd.DataFrame] = []
    stats: Dict[str, dict] = {}
    n_sheets = max(1, len(sheets))
    gpon_frames: List[pd.DataFrame] = []

    for i, (sheet_name, df) in enumerate(sheets.items()):
        orig_rows = len(df)
        df2, dropped_cols = _drop_columns(df)
        df3, closed_removed, status_col = _filter_closed_status(df2)
        cleaned_frames.append(df3)
        gp, _ = _isolate_gpon_rows(df3)
        if not gp.empty:
            gpon_frames.append(gp)
        stats[sheet_name] = {
            "original_rows": orig_rows,
            "final_rows": len(df3),
            "dropped_columns": dropped_cols,
            "status_column": status_col,
            "closed_rows_removed": closed_removed,
            "gpon_rows": len(gp),
        }
        _progress("filter", i + 1, n_sheets)

    gpon_all = pd.concat(gpon_frames, ignore_index=True) if gpon_frames else pd.DataFrame()

    _progress("classify", 0, 4)

    # 1. Pull special-status rows out of GPON
    status_subsets = _collect_status_subsets(gpon_all)
    if not gpon_all.empty:
        status_col = _find_column_case_insensitive(gpon_all, STATUS_COLUMN_NORMALIZED)
        if status_col is not None:
            norm = gpon_all[status_col].astype(str).str.strip().str.lower()
            mask_special = norm.isin(list(STATUS_SPECIAL_SHEETS.keys()))
            gpon_all = gpon_all.loc[~mask_special].reset_index(drop=True)
    _progress("classify", 1, 4)

    # 2. Detect double tickets and pull them out
    double_tickets_df, gpon_all = _detect_double_tickets(gpon_all)
    _progress("classify", 2, 4)

    # 3. Pull splitters out
    splitter_all, gpon_all = _extract_splitter_rows(gpon_all)
    _progress("classify", 3, 4)

    # 4. MDU LOS / MDU Low RX (subsets of remaining GPON, kept inline; written
    #    as separate sheets but NOT removed from per-Category 1 sheets)
    cat1_col = _find_column_case_insensitive(gpon_all, CATEGORY1_COLUMN_NORMALIZED)
    assign_col = _find_column_case_insensitive(gpon_all, ASSIGN_COLUMN_NORMALIZED)
    mdu_los_df = pd.DataFrame()
    mdu_lowrx_df = pd.DataFrame()
    if cat1_col is not None and not gpon_all.empty:
        c1n = gpon_all[cat1_col].astype(str).str.strip().str.lower()
        m_mdu_los = c1n.str.contains("mdu", na=False) & c1n.str.contains("los", na=False)
        m_mdu_lowrx = c1n.str.contains("mdu", na=False) & (c1n.str.contains("low rx", na=False) | c1n.str.contains("lowrx", na=False))
        mdu_los_df = gpon_all.loc[m_mdu_los].reset_index(drop=True)
        mdu_lowrx_df = gpon_all.loc[m_mdu_lowrx].reset_index(drop=True)

    # GPON LOS - Customer Action subset
    gpon_los_ca_df = pd.DataFrame()
    cust_action_df = status_subsets.get("Customer Action", pd.DataFrame())
    if not cust_action_df.empty:
        cat1_ca = _find_column_case_insensitive(cust_action_df, CATEGORY1_COLUMN_NORMALIZED)
        if cat1_ca is not None:
            c1ca = cust_action_df[cat1_ca].astype(str).str.strip().str.lower()
            m_los_ca = (c1ca.str.contains("los", na=False)
                        & c1ca.str.contains("fiber", na=False)
                        & c1ca.str.contains("cut", na=False)
                        & c1ca.str.contains("cable", na=False))
            gpon_los_ca_df = cust_action_df.loc[m_los_ca].reset_index(drop=True)

    _progress("classify", 4, 4)

    # ---- Write phase ----
    _progress("write", 0, 1)

    # Build "buckets" by Category 1 from the remaining GPON data
    cat1_buckets: List[Tuple[str, pd.DataFrame]] = []
    if not gpon_all.empty and cat1_col is not None:
        for cat_val, grp in gpon_all.groupby(cat1_col, dropna=False):
            cat_label = str(cat_val).strip() if pd.notna(cat_val) and str(cat_val).strip() else "(Blank)"
            sheet_base = f"GPON - {cat_label}"
            grp_sorted = grp
            if assign_col is not None and assign_col in grp.columns:
                grp_sorted = grp.sort_values(by=[assign_col], kind="stable")
            cat1_buckets.append((sheet_base, grp_sorted.reset_index(drop=True)))

    # Splitter buckets grouped by Cat1
    splitter_buckets: List[Tuple[str, pd.DataFrame]] = []
    if not splitter_all.empty:
        scat1 = _find_column_case_insensitive(splitter_all, CATEGORY1_COLUMN_NORMALIZED)
        sassign = _find_column_case_insensitive(splitter_all, ASSIGN_COLUMN_NORMALIZED)
        if scat1 is not None:
            for cat_val, grp in splitter_all.groupby(scat1, dropna=False):
                cat_label = str(cat_val).strip() if pd.notna(cat_val) and str(cat_val).strip() else "(Blank)"
                sheet_base = f"Splitter - {cat_label}"
                grp_sorted = grp
                if sassign is not None and sassign in grp.columns:
                    grp_sorted = grp.sort_values(by=[sassign], kind="stable")
                splitter_buckets.append((sheet_base, grp_sorted.reset_index(drop=True)))
        else:
            splitter_buckets.append(("Splitter", splitter_all.reset_index(drop=True)))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used: set = set()
        all_summary_blocks: List[Tuple[str, pd.DataFrame]] = []

        # Special status sheets
        for disp in ["Customer Action", "Under Monitoring", "Temporary Restoration"]:
            sd = status_subsets.get(disp, pd.DataFrame())
            if sd is None or sd.empty:
                continue
            sheet_name = _excel_safe_sheet_name(disp, used)
            sd_out = _append_sla_days(sd)
            sd_out.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], sd_out, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append((disp, sd))

        # Double Tickets
        if not double_tickets_df.empty:
            sheet_name = _excel_safe_sheet_name("Double Tickets", used)
            dt_out = _append_sla_days(double_tickets_df)
            dt_out.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], dt_out, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append(("Double Tickets", double_tickets_df))

        # GPON LOS - Customer Action
        if not gpon_los_ca_df.empty:
            sheet_name = _excel_safe_sheet_name("GPON LOS - Customer Action", used)
            df_out = _append_sla_days(gpon_los_ca_df)
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], df_out, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)

        # MDU LOS / Low RX
        if not mdu_los_df.empty:
            sheet_name = _excel_safe_sheet_name("MDU - LOS", used)
            df_out = _append_sla_days(mdu_los_df)
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], df_out, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append(("MDU - LOS", mdu_los_df))
        if not mdu_lowrx_df.empty:
            sheet_name = _excel_safe_sheet_name("MDU - Low RX", used)
            df_out = _append_sla_days(mdu_lowrx_df)
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], df_out, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append(("MDU - Low RX", mdu_lowrx_df))

        # Splitter sheets
        for sheet_base, grp_df in splitter_buckets:
            sheet_name = _excel_safe_sheet_name(sheet_base, used)
            grp_sla = _append_sla_days(grp_df)
            grp_sla.to_excel(writer, sheet_name=sheet_name, index=False)
            sass = _find_column_case_insensitive(grp_df, ASSIGN_COLUMN_NORMALIZED)
            apply_sheet_styling(writer.sheets[sheet_name], grp_sla, assign_col_name=sass,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append((sheet_base, grp_df))

        # Per-Category 1 GPON sheets
        for sheet_base, grp_df in cat1_buckets:
            sheet_name = _excel_safe_sheet_name(sheet_base, used)
            grp_sla = _append_sla_days(grp_df)
            grp_sla.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_sheet_styling(writer.sheets[sheet_name], grp_sla, assign_col_name=assign_col,
                                 sheet_label=sheet_name, settings=settings)
            all_summary_blocks.append((sheet_base, grp_df))

        # GPON SUMMARY sheet — adaptive packing
        # Build priority-ordered summary blocks:
        # Use friendly bucket labels from the data we've already separated
        summary_blocks: List[Tuple[str, pd.DataFrame]] = []

        # Pull friendly sub-buckets out of gpon_all (now without status/doubles/splitters)
        if not gpon_all.empty and cat1_col is not None:
            c1n_full = gpon_all[cat1_col].astype(str).str.strip().str.lower()
            # LOS Fiber Cut Cable
            m_los = (c1n_full.str.contains("los", na=False)
                     & c1n_full.str.contains("fiber", na=False)
                     & c1n_full.str.contains("cut", na=False)
                     & c1n_full.str.contains("cable", na=False))
            df_los = gpon_all.loc[m_los].copy()
            if not df_los.empty:
                summary_blocks.append(("GPON LOS Fiber Cut", df_los))

            m_ext = c1n_full.str.contains("extreme", na=False) & (c1n_full.str.contains("low rx", na=False) | c1n_full.str.contains("lowrx", na=False))
            df_ext = gpon_all.loc[m_ext].copy()
            if not df_ext.empty:
                summary_blocks.append(("GPON Extreme Low RX", df_ext))

            m_low = (c1n_full.str.contains("low rx", na=False) | c1n_full.str.contains("lowrx", na=False)) & (~m_ext)
            df_low = gpon_all.loc[m_low].copy()
            if not df_low.empty:
                summary_blocks.append(("GPON Low RX", df_low))

            m_losi = c1n_full.str.contains("losi", na=False)
            df_losi = gpon_all.loc[m_losi].copy()
            if not df_losi.empty:
                summary_blocks.append(("GPON LOSi", df_losi))

            m_sfi = c1n_full.str.contains("sfi", na=False)
            df_sfi = gpon_all.loc[m_sfi].copy()
            if not df_sfi.empty:
                summary_blocks.append(("GPON SFi", df_sfi))

            m_lofi = c1n_full.str.contains("lofi", na=False)
            df_lofi = gpon_all.loc[m_lofi].copy()
            if not df_lofi.empty:
                summary_blocks.append(("GPON LOFi", df_lofi))

        # Add status, double tickets, MDU
        for disp in ["Customer Action", "Under Monitoring", "Temporary Restoration"]:
            sd = status_subsets.get(disp, pd.DataFrame())
            if sd is not None and not sd.empty:
                summary_blocks.append((disp, sd))
        if not double_tickets_df.empty:
            summary_blocks.append(("Double Tickets", double_tickets_df))
        if not mdu_los_df.empty:
            summary_blocks.append(("MDU - LOS", mdu_los_df))
        if not mdu_lowrx_df.empty:
            summary_blocks.append(("MDU - Low RX", mdu_lowrx_df))

        # Splitter blocks
        for sheet_base, grp_df in splitter_buckets:
            if grp_df is not None and not grp_df.empty:
                summary_blocks.append((sheet_base, grp_df))

        # Sort by priority (most important first, then by count)
        priority_index = {name: i for i, name in enumerate(GPON_BUCKET_PRIORITY)}
        summary_blocks_sorted = sorted(
            summary_blocks,
            key=lambda kv: (priority_index.get(kv[0], 999), -len(kv[1])),
        )

        sum_name = _excel_safe_sheet_name("GPON SUMMARY", used)
        pd.DataFrame().to_excel(writer, sheet_name=sum_name, index=False)
        ws_sum = writer.sheets[sum_name]
        apply_summary_styling_adaptive(
            ws_sum,
            summary_blocks_sorted,
            assign_col_name=assign_col,
            input_path=input_path,
            report_kind="GPON",
            n_columns=4,
        )

        # Move GPON SUMMARY to the front
        try:
            wb = writer.book
            if sum_name in writer.sheets:
                ws = writer.sheets[sum_name]
                wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))
        except Exception:
            pass

    _progress("done", 1, 1)
    return stats


# ============================================================================
# SECTION 9   GUI
# ============================================================================
#
# One window, two tabs.  Each tab has:
#     - drag-and-drop input (or browse)
#     - live preview showing what will be generated
#     - "Run" button with progress bar
#     - status / log area
# Plus shared:
#     - File menu with Recent files
#     - Tools menu with Settings dialog
#     - Help menu


def load_settings() -> Dict[str, object]:
    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE) as f:
                data = json.load(f)
            merged = dict(DEFAULT_SETTINGS)
            merged.update(data)
            return merged
        except Exception:
            return dict(DEFAULT_SETTINGS)
    return dict(DEFAULT_SETTINGS)


def save_settings(settings: Dict[str, object]) -> None:
    try:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(settings, f, indent=2)
    except Exception:
        pass


def _date_title() -> str:
    """Friendly suffixed date like '28th April 2026'."""
    dt = datetime.now()
    d = dt.day
    if 10 <= (d % 100) <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(d % 10, "th")
    return f"{d}{suf} {dt.strftime('%B')} {dt.year}"


# ---- Bucket-count chart widget (pure tkinter Canvas) ----
def _make_bucket_chart(parent, height=240):
    """Returns a tk.Canvas configured for horizontal bar charts of bucket counts."""
    canvas = tk.Canvas(parent, bg="white", highlightthickness=0, height=height)
    canvas._counts = {}  # type: ignore
    canvas._color_map = {}  # type: ignore

    BUCKET_COLORS = {
        # Enterprise
        "Equipment TTs": "#7030A0",
        "Power": "#C00000",
        "BB LOS": "#4472C4",
        "BB on Extreme Low RX": "#4472C4",
        "BB on Low RX": "#4472C4",
        "BB on CRC Errors": "#4472C4",
        "6150|6120|1050-LOS": "#ED7D31",
        "6150|6120|1050-Extreme Low RX": "#ED7D31",
        "6150|6120|1050-Low RX": "#ED7D31",
        "6150|6120|1050-CRC": "#ED7D31",
        "Access Ring-LOS": "#70AD47",
        "Access Ring-Low RX": "#70AD47",
        "Access Ring-Offline": "#70AD47",
        # GPON
        "Customer Action": "#ED7D31",
        "Under Monitoring": "#70AD47",
        "Temporary Restoration": "#4472C4",
        "Double Tickets": "#C00000",
    }
    canvas._color_map = BUCKET_COLORS  # type: ignore

    def redraw():
        canvas.delete("all")
        w = canvas.winfo_width()
        h = canvas.winfo_height()
        if w < 50 or h < 50:
            return
        if not canvas._counts:  # type: ignore
            canvas.create_text(w / 2, h / 2,
                               text="Load an input file to preview",
                               fill="#888888",
                               font=("TkDefaultFont", 10, "italic"))
            return
        items = [(k, v) for k, v in canvas._counts.items() if v > 0]  # type: ignore
        if not items:
            canvas.create_text(w / 2, h / 2,
                               text="No tickets matched any bucket",
                               fill="#888888",
                               font=("TkDefaultFont", 10, "italic"))
            return
        items.sort(key=lambda kv: kv[1], reverse=True)
        max_v = max(v for _, v in items)
        label_w = 180
        count_w = 50
        pad_x = 12
        pad_y = 8
        bar_left = pad_x + label_w + 4
        bar_right = w - pad_x - count_w - 4
        bar_w = max(20, bar_right - bar_left)
        n = len(items)
        bar_h = max(14, min(28, (h - 2 * pad_y) / max(n, 1) - 4))
        gap = 4
        y = pad_y
        for label, val in items:
            color = canvas._color_map.get(label, "#808080")  # type: ignore
            display_label = label if len(label) <= 28 else label[:26] + "\u2026"
            canvas.create_text(pad_x + label_w, y + bar_h / 2,
                               text=display_label, anchor="e",
                               font=("TkDefaultFont", 9))
            wpx = bar_w * (val / max_v) if max_v else 0
            if wpx >= 1:
                canvas.create_rectangle(bar_left, y, bar_left + wpx, y + bar_h,
                                        fill=color, outline="")
            canvas.create_text(bar_right + 4, y + bar_h / 2,
                               text=str(val), anchor="w",
                               font=("TkDefaultFont", 9, "bold"))
            y += bar_h + gap

    def set_counts(counts):
        canvas._counts = dict(counts)  # type: ignore
        redraw()

    def clear():
        canvas._counts = {}  # type: ignore
        redraw()

    canvas.set_counts = set_counts  # type: ignore
    canvas.clear = clear  # type: ignore
    canvas.bind("<Configure>", lambda _e: redraw())
    return canvas


class SettingsDialog(tk.Toplevel):
    """Modal dialog for SLA thresholds and ttk theme."""

    def __init__(self, parent, settings, on_save):
        super().__init__(parent)
        self.title("Settings")
        self.geometry("440x360")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self._settings = dict(settings)
        self._on_save = on_save

        body = ttk.Frame(self, padding=14)
        body.pack(fill=tk.BOTH, expand=True)

        ttk.Label(body, text="SLA color thresholds (days)",
                  font=("TkDefaultFont", 10, "bold")).grid(row=0, column=0, columnspan=2, sticky="w")

        ttk.Label(body, text="Green \u2264").grid(row=1, column=0, sticky="w", pady=(8, 4))
        self.var_green = tk.IntVar(value=int(settings.get("sla_green_max", 1)))
        ttk.Spinbox(body, from_=0, to=30, textvariable=self.var_green, width=6).grid(row=1, column=1, sticky="w", pady=(8, 4))

        ttk.Label(body, text="Yellow \u2264").grid(row=2, column=0, sticky="w", pady=4)
        self.var_yellow = tk.IntVar(value=int(settings.get("sla_yellow_max", 3)))
        ttk.Spinbox(body, from_=0, to=60, textvariable=self.var_yellow, width=6).grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(body, text="Orange \u2264").grid(row=3, column=0, sticky="w", pady=4)
        self.var_orange = tk.IntVar(value=int(settings.get("sla_orange_max", 7)))
        ttk.Spinbox(body, from_=0, to=90, textvariable=self.var_orange, width=6).grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(body, text="Red \u2264").grid(row=4, column=0, sticky="w", pady=4)
        self.var_red = tk.IntVar(value=int(settings.get("sla_red_max", 14)))
        ttk.Spinbox(body, from_=0, to=180, textvariable=self.var_red, width=6).grid(row=4, column=1, sticky="w", pady=4)

        ttk.Label(body, text="(Anything above the Red threshold is shown as critical / dark red.)",
                  font=("TkDefaultFont", 9), foreground="#666666",
                  wraplength=400).grid(row=5, column=0, columnspan=2, sticky="w", pady=(2, 12))

        ttk.Separator(body).grid(row=6, column=0, columnspan=2, sticky="we", pady=(0, 10))

        ttk.Label(body, text="Theme",
                  font=("TkDefaultFont", 10, "bold")).grid(row=7, column=0, columnspan=2, sticky="w")
        ttk.Label(body, text="ttk style:").grid(row=8, column=0, sticky="w", pady=(8, 4))
        self.var_theme = tk.StringVar(value=str(settings.get("theme", "clam")))
        themes = list(ttk.Style().theme_names())
        ttk.Combobox(body, textvariable=self.var_theme, values=themes,
                     state="readonly", width=14).grid(row=8, column=1, sticky="w", pady=(8, 4))

        btn_row = ttk.Frame(body)
        btn_row.grid(row=10, column=0, columnspan=2, sticky="e", pady=(18, 0))
        ttk.Button(btn_row, text="Cancel", command=self.destroy).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_row, text="Save", command=self._save).pack(side=tk.LEFT)

        body.grid_columnconfigure(0, weight=1)

    def _save(self):
        g = self.var_green.get()
        y = self.var_yellow.get()
        o = self.var_orange.get()
        r = self.var_red.get()
        if not (g < y < o < r):
            messagebox.showerror("Invalid thresholds",
                                 "Thresholds must be strictly ascending: Green < Yellow < Orange < Red.",
                                 parent=self)
            return
        self._settings["sla_green_max"] = g
        self._settings["sla_yellow_max"] = y
        self._settings["sla_orange_max"] = o
        self._settings["sla_red_max"] = r
        self._settings["theme"] = self.var_theme.get()
        self._on_save(self._settings)
        self.destroy()


class ReportTab:
    """A single tab handling either GPON or Enterprise report generation.

    Each tab is independent — own input/output paths, own preview, own log.
    Shares settings + recent-files with the parent app.
    """

    def __init__(self, parent_frame, app, kind: str):
        """kind: 'gpon' or 'enterprise'."""
        self.frame = parent_frame
        self.app = app
        self.kind = kind  # 'gpon' or 'enterprise'
        self.is_gpon = (kind == "gpon")

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")

        self._build_ui()
        self.input_var.trace_add("write", self._on_input_changed)

    @property
    def report_label(self) -> str:
        return "GPON" if self.is_gpon else "Enterprise"

    @property
    def recent_key(self) -> str:
        return "recent_gpon_files" if self.is_gpon else "recent_ent_files"

    @property
    def preview_fn(self):
        return gpon_preview if self.is_gpon else enterprise_preview

    @property
    def process_fn(self):
        return gpon_process if self.is_gpon else enterprise_process

    def _build_ui(self):
        outer = ttk.Frame(self.frame, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        # Drop zone + file selection
        drop_section = ttk.Labelframe(outer, text=f"  {self.report_label} input file  ",
                                        padding=10)
        drop_section.pack(fill=tk.X, pady=(0, 8))

        if _DND_AVAILABLE:
            drop_text = "Drop the Issue.xlsx file here, or click Browse\u2026"
        else:
            drop_text = "Click Browse to select Issue.xlsx (install tkinterdnd2 for drag-and-drop)"
        self.drop_label = tk.Label(drop_section, text=drop_text,
                                    bg="#F2F6FC", fg="#1F3864",
                                    relief="solid", borderwidth=1, padx=12, pady=12,
                                    font=("TkDefaultFont", 11))
        self.drop_label.pack(fill=tk.X, pady=(0, 8))
        if _DND_AVAILABLE:
            try:
                self.drop_label.drop_target_register(DND_FILES)
                self.drop_label.dnd_bind("<<Drop>>", self._on_drop)
            except Exception:
                pass

        row1 = ttk.Frame(drop_section)
        row1.pack(fill=tk.X)
        ttk.Label(row1, text="Input:").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Entry(row1, textvariable=self.input_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(row1, text="Browse\u2026", command=self._browse_input).pack(side=tk.LEFT, padx=(6, 0))

        row2 = ttk.Frame(drop_section)
        row2.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(row2, text="Output:").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Entry(row2, textvariable=self.output_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(row2, text="Save As\u2026", command=self._browse_output).pack(side=tk.LEFT, padx=(6, 0))

        # Middle: live preview + chart
        mid = ttk.Frame(outer)
        mid.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        preview_frame = ttk.Labelframe(mid, text="  Live preview  ", padding=10)
        preview_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))

        self.kpi_frame = ttk.Frame(preview_frame)
        self.kpi_frame.pack(fill=tk.X, pady=(0, 8))

        # Different KPIs per tab
        if self.is_gpon:
            kpi_specs = [
                ("Total rows", "kpi_total"),
                ("Dropped (closed)", "kpi_closed"),
                ("GPON rows", "kpi_gpon"),
                ("Doubles", "kpi_doubles"),
                ("Splitters", "kpi_splitters"),
            ]
        else:
            kpi_specs = [
                ("Total rows", "kpi_total"),
                ("Dropped (closed)", "kpi_closed"),
                ("Dropped (GPON)", "kpi_gpon"),
                ("Will process", "kpi_kept"),
                ("CESR matches", "kpi_cesr"),
            ]

        self._kpi_widgets = {}
        for col, (label, attr) in enumerate(kpi_specs):
            self._kpi_widgets[attr] = self._kpi_card(self.kpi_frame, label, "—", col)
            self.kpi_frame.grid_columnconfigure(col, weight=1, uniform="kpi")

        ttk.Label(preview_frame, text="Source sheet info:",
                  font=("TkDefaultFont", 10, "bold")).pack(anchor="w")
        self.txt_info = scrolledtext.ScrolledText(preview_frame, height=8, wrap=tk.WORD,
                                                   font=("TkFixedFont", 9))
        self.txt_info.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.txt_info.config(state=tk.DISABLED)

        chart_frame = ttk.Labelframe(mid, text="  Distribution preview  ", padding=10)
        chart_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6, 0))
        self.chart = _make_bucket_chart(chart_frame, height=320)
        self.chart.pack(fill=tk.BOTH, expand=True)

        # Action bar
        action = ttk.Frame(outer)
        action.pack(fill=tk.X, pady=(0, 6))
        self.btn_run = ttk.Button(action, text=f"Generate {self.report_label} report",
                                   command=self._run_clicked, state=tk.DISABLED)
        self.btn_run.pack(side=tk.LEFT)
        self.btn_open = ttk.Button(action, text="Open output", command=self._open_output, state=tk.DISABLED)
        self.btn_open.pack(side=tk.LEFT, padx=(6, 0))
        self.btn_open_folder = ttk.Button(action, text="Show in folder",
                                           command=self._open_output_folder, state=tk.DISABLED)
        self.btn_open_folder.pack(side=tk.LEFT, padx=(6, 0))
        self.progress = ttk.Progressbar(action, mode="determinate", length=220)
        self.progress.pack(side=tk.RIGHT)

        statusbar = ttk.Frame(outer, relief="sunken")
        statusbar.pack(fill=tk.X)
        ttk.Label(statusbar, textvariable=self.status_var, anchor="w",
                  padding=(8, 4)).pack(fill=tk.X)

    def _kpi_card(self, parent, label, value, col):
        card = ttk.Frame(parent, relief="solid", borderwidth=1, padding=8)
        card.grid(row=0, column=col, sticky="nsew", padx=2)
        ttk.Label(card, text=label.upper(),
                  font=("TkDefaultFont", 8, "bold"), foreground="#666666").pack()
        val = ttk.Label(card, text=value,
                        font=("TkDefaultFont", 16, "bold"), foreground="#1F3864")
        val.pack()
        return val

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title=f"Select the {self.report_label} input file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self._load_input(path)

    def _on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        path = raw.split("} {")[0].strip("{}")
        if path and Path(path).exists():
            self._load_input(path)

    def _load_input(self, path: str):
        if not Path(path).exists():
            messagebox.showerror("File not found", f"Could not find:\n{path}")
            return
        self.input_var.set(path)

    def _default_output_path(self, input_path: Path) -> Path:
        last_dir = self.app.settings.get("last_output_dir") or ""
        out_name = f"{_date_title()} {self.report_label} Report.xlsx"
        if last_dir and Path(last_dir).is_dir():
            return Path(last_dir) / out_name
        return input_path.with_name(out_name)

    def _browse_output(self):
        suggested = self.output_var.get().strip()
        if not suggested:
            in_path = self.input_var.get().strip()
            if in_path:
                suggested = str(self._default_output_path(Path(in_path)))
        initialdir = None
        initialfile = None
        if suggested:
            sp = Path(suggested)
            initialdir = str(sp.parent)
            initialfile = sp.name
        path = filedialog.asksaveasfilename(
            title=f"Save {self.report_label} report as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=initialdir,
            initialfile=initialfile,
        )
        if path:
            if not str(path).lower().endswith(".xlsx"):
                path = str(path) + ".xlsx"
            self.output_var.set(path)

    def _on_input_changed(self, *_):
        in_path = self.input_var.get().strip()
        if not in_path or not Path(in_path).exists():
            self.output_var.set("")
            self.btn_run.config(state=tk.DISABLED)
            self.chart.clear()
            self._set_kpis_empty()
            self._set_info("")
            return
        self.output_var.set(str(self._default_output_path(Path(in_path))))
        self.btn_run.config(state=tk.NORMAL)
        self.btn_open.config(state=tk.DISABLED)
        self.btn_open_folder.config(state=tk.DISABLED)
        self._start_preview(Path(in_path))

    def _start_preview(self, path: Path):
        self.status_var.set("Reading preview\u2026")
        self._set_kpis_empty()
        self.chart.clear()
        self._set_info(f"Loading {path.name}\u2026")
        t = threading.Thread(target=self._preview_worker, args=(path,), daemon=True)
        t.start()

    def _preview_worker(self, path: Path):
        try:
            info = self.preview_fn(path)
        except Exception as exc:
            err = f"Preview failed: {exc}\n{traceback.format_exc()}"
            self.app.root.after(0, lambda: self._preview_error(err))
            return
        self.app.root.after(0, lambda: self._preview_ok(path, info))

    def _preview_error(self, msg):
        self.status_var.set("Preview failed")
        self._set_info(msg)
        self.chart.clear()

    def _preview_ok(self, path: Path, info: Dict[str, object]):
        self._set_kpis(info)
        # Build chart counts from the appropriate field
        if self.is_gpon:
            counts = dict(info.get("status_counts", {}))
            counts["Double Tickets"] = info.get("double_tickets", 0)
            counts["Splitter rows"] = info.get("splitter_rows", 0)
            self.chart.set_counts(counts)
        else:
            self.chart.set_counts(info.get("bucket_counts", {}))

        # Detail text
        lines: List[str] = []
        lines.append(f"File:        {path.name}")
        lines.append(f"Sheets:      {', '.join(info.get('sheet_names', []))}")
        lines.append(f"Total rows:  {info.get('total_rows', 0)}")
        lines.append(f"Cleaned:     {info.get('cleaned_total', 0)} (dropped closed: {info.get('dropped_closed', 0)})")
        if self.is_gpon:
            lines.append(f"GPON rows:   {info.get('gpon_rows', 0)}  (non-GPON: {info.get('non_gpon_rows', 0)})")
            lines.append(f"Doubles:     {info.get('double_tickets', 0)}")
            lines.append(f"Splitters:   {info.get('splitter_rows', 0)}")
            lines.append("")
            lines.append("Status breakdown (GPON only):")
            for k, v in info.get("status_counts", {}).items():
                lines.append(f"  {k:<24} {v}")
            lines.append("")
            lines.append("Top Category 1 buckets:")
            cats = info.get("category_counts", {})
            for k, v in sorted(cats.items(), key=lambda kv: -kv[1])[:8]:
                lines.append(f"  {k[:55]:<55} {v}")
        else:
            lines.append(f"After GPON exclusion: {info.get('kept_rows', 0)}")
            lines.append("")
            lines.append("Bucket distribution:")
            buckets = info.get("bucket_counts", {})
            non_zero = sorted(((k, v) for k, v in buckets.items() if v > 0),
                              key=lambda kv: kv[1], reverse=True)
            if not non_zero:
                lines.append("  (no rows matched any bucket)")
            for k, v in non_zero:
                lines.append(f"  {k:<32} {v}")
            cesr = info.get("cesr_count", 0)
            if cesr:
                lines.append(f"  CESR (extra sheet)              {cesr}")
        self._set_info("\n".join(lines))
        self.status_var.set("Preview ready \u2014 click Generate when you're set.")

    def _set_info(self, text: str):
        self.txt_info.config(state=tk.NORMAL)
        self.txt_info.delete("1.0", tk.END)
        self.txt_info.insert(tk.END, text)
        self.txt_info.config(state=tk.DISABLED)

    def _set_kpis_empty(self):
        for w in self._kpi_widgets.values():
            w.config(text="—")

    def _set_kpis(self, info: Dict[str, object]):
        if self.is_gpon:
            self._kpi_widgets["kpi_total"].config(text=str(info.get("total_rows", 0)))
            self._kpi_widgets["kpi_closed"].config(text=str(info.get("dropped_closed", 0)))
            self._kpi_widgets["kpi_gpon"].config(text=str(info.get("gpon_rows", 0)))
            self._kpi_widgets["kpi_doubles"].config(text=str(info.get("double_tickets", 0)))
            self._kpi_widgets["kpi_splitters"].config(text=str(info.get("splitter_rows", 0)))
        else:
            self._kpi_widgets["kpi_total"].config(text=str(info.get("total_rows", 0)))
            self._kpi_widgets["kpi_closed"].config(text=str(info.get("dropped_closed", 0)))
            self._kpi_widgets["kpi_gpon"].config(text=str(info.get("dropped_gpon", 0)))
            self._kpi_widgets["kpi_kept"].config(text=str(info.get("kept_rows", 0)))
            self._kpi_widgets["kpi_cesr"].config(text=str(info.get("cesr_count", 0)))

    def _run_clicked(self):
        in_path = self.input_var.get().strip()
        out_path = self.output_var.get().strip()
        if not in_path or not Path(in_path).exists():
            messagebox.showerror("No input", "Choose a valid input file first.")
            return
        if not out_path:
            messagebox.showerror("No output", "Choose where to save the output.")
            return
        self.btn_run.config(state=tk.DISABLED)
        self.btn_open.config(state=tk.DISABLED)
        self.btn_open_folder.config(state=tk.DISABLED)
        self.progress.config(value=0, maximum=100)
        self.status_var.set("Generating report\u2026")
        self.app.push_recent(self.recent_key, in_path)
        self.app.settings["last_output_dir"] = str(Path(out_path).parent)
        save_settings(self.app.settings)

        t = threading.Thread(target=self._run_worker,
                             args=(Path(in_path), Path(out_path)), daemon=True)
        t.start()

    def _run_worker(self, in_path: Path, out_path: Path):
        phase_pct = {
            "load":     10,
            "filter":   30,
            "classify": 55,
            "write":    75,
            "done":    100,
        }

        def cb(phase, current, total):
            pct = phase_pct.get(phase, 50)
            if phase == "filter" and total > 0:
                pct = 10 + int(20 * current / total)
            elif phase == "classify" and total > 0:
                pct = 30 + int(25 * current / total)
            self.app.root.after(0, lambda: self.progress.config(value=pct))
            self.app.root.after(0, lambda: self.status_var.set(f"Working \u2014 {phase}\u2026"))

        try:
            stats = self.process_fn(in_path, out_path,
                                     progress_callback=cb,
                                     settings=self.app.settings)
        except Exception as exc:
            tb = traceback.format_exc()
            self.app.root.after(0, lambda: self._run_error(exc, tb))
            return
        self.app.root.after(0, lambda: self._run_ok(out_path, stats))

    def _run_error(self, exc, tb):
        self.progress.config(value=0)
        self.status_var.set("Failed")
        self.btn_run.config(state=tk.NORMAL)
        messagebox.showerror("Generate failed", f"{exc}\n\n{tb}")

    def _run_ok(self, out_path: Path, stats):
        self.progress.config(value=100)
        n_rows = sum(s.get("final_rows", 0) for s in stats.values())
        self.status_var.set(f"Done \u2014 {n_rows} rows written to {out_path.name}")
        self.btn_run.config(state=tk.NORMAL)
        self.btn_open.config(state=tk.NORMAL)
        self.btn_open_folder.config(state=tk.NORMAL)
        messagebox.showinfo("Report generated",
                            f"Wrote: {out_path.name}\n\nLocation:\n{out_path.parent}")

    def _open_output(self):
        out = self.output_var.get().strip()
        if not out:
            return
        p = Path(out)
        if not p.exists():
            messagebox.showwarning("Open output", f"File not found:\n{p}")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(p))  # noqa
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", str(p)])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", str(p)])
        except Exception as e:
            messagebox.showerror("Open failed", str(e))

    def _open_output_folder(self):
        out = self.output_var.get().strip()
        if not out:
            return
        folder = Path(out).parent
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(folder))  # noqa
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", str(folder)])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", str(folder)])
        except Exception as e:
            messagebox.showerror("Open folder failed", str(e))


class ReportApp:
    """Main window: notebook with two tabs (GPON | Enterprise)."""
    APP_TITLE = "Report Gen \u2014 JTL NOC Monitoring"

    def __init__(self, root):
        self.root = root
        self.settings = load_settings()
        self.root.title(self.APP_TITLE)
        self.root.geometry("1080x720")
        self.root.minsize(960, 640)

        try:
            ttk.Style().theme_use(self.settings.get("theme", "clam"))
        except tk.TclError:
            pass

        self._setup_styles()
        self._build_ui()
        self._refresh_recent_menus()

    def _setup_styles(self):
        s = ttk.Style()
        s.configure("Header.TLabel", font=("TkDefaultFont", 14, "bold"), foreground="#1F3864")
        s.configure("Sub.TLabel", font=("TkDefaultFont", 10), foreground="#666666")
        s.configure("Run.TButton", font=("TkDefaultFont", 11, "bold"))

    def _build_ui(self):
        # Menu bar
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        self.recent_gpon_menu = tk.Menu(file_menu, tearoff=0)
        self.recent_ent_menu = tk.Menu(file_menu, tearoff=0)
        file_menu.add_cascade(label="Recent GPON files", menu=self.recent_gpon_menu)
        file_menu.add_cascade(label="Recent Enterprise files", menu=self.recent_ent_menu)
        file_menu.add_separator()
        file_menu.add_command(label="Quit", command=self.root.destroy)
        menubar.add_cascade(label="File", menu=file_menu)

        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="Settings\u2026", command=self._open_settings)
        menubar.add_cascade(label="Tools", menu=tools_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self._about)
        menubar.add_cascade(label="Help", menu=help_menu)
        self.root.config(menu=menubar)

        # Title bar
        top = ttk.Frame(self.root, padding=(12, 12, 12, 0))
        top.pack(fill=tk.X)
        ttk.Label(top, text="Report Gen", style="Header.TLabel").pack(side=tk.LEFT)
        ttk.Label(top, text=f"v{APP_VERSION}  \u2022  \u00a9 {APP_AUTHOR}",
                  style="Sub.TLabel").pack(side=tk.RIGHT)

        # Notebook with two tabs
        nb = ttk.Notebook(self.root)
        nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
        self.tab_gpon_frame = ttk.Frame(nb)
        self.tab_ent_frame = ttk.Frame(nb)
        nb.add(self.tab_gpon_frame, text="GPON")
        nb.add(self.tab_ent_frame, text="Enterprise")

        self.gpon_tab = ReportTab(self.tab_gpon_frame, self, "gpon")
        self.ent_tab = ReportTab(self.tab_ent_frame, self, "enterprise")

    def push_recent(self, key: str, path: str):
        recent = list(self.settings.get(key, []) or [])
        if path in recent:
            recent.remove(path)
        recent.insert(0, path)
        recent = recent[:5]
        self.settings[key] = recent
        save_settings(self.settings)
        self._refresh_recent_menus()

    def _refresh_recent_menus(self):
        for menu, key, tab in [
            (self.recent_gpon_menu, "recent_gpon_files", self.gpon_tab),
            (self.recent_ent_menu, "recent_ent_files", self.ent_tab),
        ]:
            menu.delete(0, tk.END)
            recent = self.settings.get(key, []) or []
            if not recent:
                menu.add_command(label="(none)", state=tk.DISABLED)
                continue
            for path in recent[:5]:
                display = path if len(path) <= 60 else "\u2026" + path[-57:]
                menu.add_command(
                    label=display,
                    command=lambda p=path, t=tab: t._load_input(p),
                )

    def _open_settings(self):
        SettingsDialog(self.root, self.settings, on_save=self._on_settings_saved)

    def _on_settings_saved(self, new_settings):
        self.settings = new_settings
        save_settings(self.settings)
        try:
            ttk.Style().theme_use(self.settings.get("theme", "clam"))
        except tk.TclError:
            pass
        self.gpon_tab.status_var.set("Settings saved.")
        self.ent_tab.status_var.set("Settings saved.")

    def _about(self):
        msg = (
            f"{self.APP_TITLE}\n"
            f"Version {APP_VERSION}\n"
            f"\u00a9 {APP_AUTHOR}\n\n"
            f"Drag-and-drop: {'enabled' if _DND_AVAILABLE else 'unavailable (install tkinterdnd2)'}\n"
        )
        messagebox.showinfo("About", msg)


# ============================================================================
# SECTION 10  Entry point
# ============================================================================

def main():
    if not _TK_AVAILABLE:
        print("tkinter is not available on this system.", file=sys.stderr)
        print("On Linux: sudo apt-get install python3-tk", file=sys.stderr)
        sys.exit(1)
    if _DND_AVAILABLE:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    ReportApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
